import os
import sys
import json
import argparse
import logging
import re
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openai import OpenAI
from dotenv import load_dotenv
import pandas as pd

# Setup Logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Styling
_FONT = Font(name='Arial', size=10)
_FILL_LLM = PatternFill('solid', start_color='DDEBF7')  # Light blue
_CENTER = Alignment(horizontal='center', vertical='center')
_LEFT = Alignment(horizontal='left', vertical='center')

def canonical_coverage_tier(value) -> str:
    """Normalize common coverage-tier aliases."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    raw = str(value).strip().lower()
    if not raw or raw in ('n/a', 'na', 'none', 'null', 'nan', '-', ''):
        return ""
    token = re.sub(r"[^a-z0-9]", "", raw).upper()
    tier_map = {
        "E": "EE", "EE": "EE", "EMPLOYEE": "EE", "EMPLOYER": "EE", "EMPLOYEEONLY": "EE",
        "S": "ES", "ES": "ES", "SPOUSE": "ES", "SS": "ES", "EMPLOYEESPOUSE": "ES", "EMPLOYEEANDSPOUSE": "ES",
        "C": "EC", "EC": "EC", "CH": "EC", "CHILD": "EC", "CHILDREN": "EC", "EMPLOYEECHILDREN": "EC", "EMPLOYEEANDCHILDREN": "EC",
        "F": "FAM", "FAM": "FAM", "FAMILY": "FAM", "EF": "FAM", "EMPLOYEEFAMILY": "FAM",
    }
    result = tier_map.get(token)
    if result: return result
    if 'spouse' in raw and ('child' in raw or 'fam' in raw or '1+' in raw): return "FAM"
    if 'spouse' in raw or 'partner' in raw: return "ES"
    if 'child' in raw or '1+' in raw or 'dep' in raw: return "EC"
    if 'only' in raw or 'employee' in raw: return "EE"
    return token

def run_llm_resolution(
    validated_excel: Path,
    audit_json: Path,
    output_excel: Path = None,
    template_type: str = 'type3',   # NEW — determines if CH/SP dependent-skip applies
) -> dict:
    if not validated_excel.exists() or not audit_json.exists():
        logger.error("Missing input files for LLM resolution.")
        return {}

    with open(audit_json, 'r', encoding='utf-8') as f:
        audit_data = json.load(f)

    entries = audit_data.get('entries', [])
    unclaimed = audit_data.get('unclaimed_invoices', [])

    # Find unresolved census rows
    unresolved_census = []
    for entry in entries:
        if entry.get('action') in ['unresolved', 'flagged_possible']:
            # We skip 'deleted_duplicate' and 'kept_unresolved_appended' which are appended rows
            unresolved_census.append({
                'row': entry['row'],
                'raw_name': entry['raw_name']
            })

    if not unresolved_census or not unclaimed:
        logger.info("No unresolved census names or unclaimed invoices to process via LLM.")
        if output_excel is None:
            output_excel = validated_excel.with_name(validated_excel.name.replace("VALIDATED_", "LLM_RESOLVED_"))
        import shutil
        shutil.copy2(str(validated_excel), str(output_excel))
        logger.info(f"Saved (skipped) LLM output to {output_excel}")
        return {'status': 'skipped', 'matches': 0, 'output_path': str(output_excel)}

    logger.info(f"LLM Resolution: {len(unresolved_census)} unresolved census names vs {len(unclaimed)} unclaimed invoices.")

    load_dotenv()
    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        logger.error("OPENAI_API_KEY missing. Skipping LLM resolution.")
        return {}

    client = OpenAI(api_key=api_key)

    # Prepare Prompt
    system_prompt = (
        "You are an expert data matching assistant. Your job is to match unresolved names from a Census "
        "to available Unclaimed names from an Invoice.\n"
        "Be logical. Look for nicknames (e.g. Robert = Bob), severe typos, or swapped names.\n"
        "DO NOT guess wildly. The first and last names must reasonably correspond. DO NOT match completely different names just because they are left over.\n"
        "Return ONLY valid JSON in this exact format:\n"
        "{\n"
        "  \"matches\": [\n"
        "    {\"census_name\": \"...\", \"invoice_name\": \"...\"}\n"
        "  ]\n"
        "}\n"
        "If a census name cannot be confidently matched to any invoice name, do not include it in the output array."
    )

    user_prompt = (
        f"Unresolved Census Names:\n{json.dumps([c['raw_name'] for c in unresolved_census], indent=2)}\n\n"
        f"Unclaimed Invoice Names:\n{json.dumps([u['raw_name'] for u in unclaimed], indent=2)}"
    )

    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            temperature=0.1
        )
        content = response.choices[0].message.content
        match_data = json.loads(content)
        matches = match_data.get('matches', [])
    except Exception as e:
        logger.error(f"OpenAI API call failed: {e}")
        return {}

    if not matches:
        logger.info("LLM found no confident matches.")
        # Still write the output file so the orchestrator's file-existence check
        # passes and does not incorrectly raise [WARN Phase 4 failed or skipped].
        # wb is not loaded yet here, so just copy the validated Excel unchanged.
        if output_excel is None:
            output_excel = validated_excel.with_name(validated_excel.name.replace("VALIDATED_", "LLM_RESOLVED_"))
        import shutil
        shutil.copy2(str(validated_excel), str(output_excel))
        logger.info(f"Saved (no-match) LLM output to {output_excel}")
        return {'status': 'completed', 'matches': 0, 'output_path': str(output_excel)}

    logger.info(f"LLM proposed {len(matches)} matches.")

    # Build lookup maps
    invoice_map = {u['raw_name'].lower(): u for u in unclaimed}
    
    # Update Excel
    wb = load_workbook(str(validated_excel))
    ws = next(
        (wb[s] for s in wb.sheetnames
         if any(k in s.lower() for k in ('census', 'employee', 'table', 'sheet'))),
        wb.active
    )

    # Find columns
    plan_col, prem_col, disc_col, rel_col, cov_col = None, None, None, None, None
    name_col, first_col, last_col = None, None, None
    for r in range(1, 40):
        for c in range(1, min(ws.max_column + 1, 60)):
            val = str(ws.cell(row=r, column=c).value or '').strip().lower()
            if 'plan' in val and 'premium' not in val and not plan_col: plan_col = c
            if 'premium' in val and not prem_col:                       prem_col = c
            if 'discrep' in val and not disc_col:                       disc_col = c
            if 'relation' in val and 'discrep' not in val and not rel_col: rel_col = c
            if ('coverage' in val or 'tier' in val) and not cov_col:     cov_col = c
            if (('employee' in val and 'name' in val) or ('full' in val and 'name' in val)) and not name_col:
                name_col = c
            if 'first' in val and 'name' in val and not first_col:
                first_col = c
            if 'last' in val and 'name' in val and not last_col:
                last_col = c
        if plan_col and prem_col and disc_col:
            break

    # Dependent relations — skip CH/SP rows for ANY template type that has
    # a Relationship column. If no such column exists, rel_col stays None
    # and the guard below never fires.
    _DEPENDENT_RELATIONS = {'ch', 'sp', 'child', 'spouse', 'dependent', 'dep'}

    match_count = 0
    target_invoice_raw_names = set()
    
    # --- NEW LOGIC: Identify all 'WO' rows directly from the Excel ---
    waiver_rows_indices = set()
    if cov_col is not None:
        for r_idx in range(1, ws.max_row + 1): # Iterate through all rows in the worksheet
            cov_val = str(ws.cell(row=r_idx, column=cov_col).value or '').strip().upper()
            if cov_val == 'WO':
                waiver_rows_indices.add(r_idx)
                logger.info(f"Identified row {r_idx} with coverage='WO' (will skip LLM filling).")
    # --- END NEW LOGIC ---

    for match in matches:
        c_name = match.get('census_name')
        i_name = match.get('invoice_name')
        if not c_name or not i_name: continue

        c_name_lower = str(c_name).strip().lower()
        target_row = next((c['row'] for c in unresolved_census if str(c['raw_name']).strip().lower() == c_name_lower), None)
        target_invoice = invoice_map.get(str(i_name).strip().lower())

        if target_row and target_invoice and disc_col:
            # ── Guard: never fill a dependent (CH / SP) row ──────────────
            if rel_col is not None:
                rel_val = str(ws.cell(row=target_row, column=rel_col).value or '').strip().lower()
                if rel_val in _DEPENDENT_RELATIONS:
                    logger.warning(
                        f"LLM proposed match for dependent row {target_row} "
                        f"(relation='{rel_val.upper()}', name='{c_name}') — SKIPPED."
                    )
                    continue

            # --- NEW LOGIC: Skip LLM filling for 'WO' rows ---
            if target_row in waiver_rows_indices:
                logger.warning(
                    f"LLM proposed match for waiver row {target_row} "
                    f"(coverage='WO', name='{c_name}') — SKIPPED LLM FILLING."
                )
                continue
            # --- END NEW LOGIC ---

            logger.info(f"Applying match: Row {target_row} ({c_name}) -> Invoice ({target_invoice['raw_name']})")
            if plan_col and target_invoice.get('plan'):
                cell = ws.cell(row=target_row, column=plan_col)
                existing_val = str(cell.value or '').strip().lower()
                is_generic = existing_val in {'base', 'core', 'buy up', 'buy-up', 'waived', 'not eligible', 'high', 'low', 'standard', 'premium', 'basic'}
                is_empty = existing_val == '' or existing_val == 'nan' or existing_val == 'none'
                
                if is_empty or is_generic:
                    cell.value = target_invoice['plan']
                    cell.font = _FONT
                    cell.alignment = _LEFT

            if prem_col and target_invoice.get('premium') is not None:
                cell = ws.cell(row=target_row, column=prem_col)
                val = cell.value
                is_empty = val is None or str(val).strip() == '' or str(val).strip().upper() in ('#N/A', 'N/A', 'NA', '0', '0.0', '0.00')
                if is_empty:
                    cell.value = target_invoice['premium']
                    cell.font = _FONT
                    cell.alignment = _CENTER
                    cell.number_format = '$#,##0.00'

            # ── Determine Coverage Status ────────────────────────────────
            cov_status = "not found on invoice"
            inv_tier = canonical_coverage_tier(target_invoice.get('coverage'))
            cen_tier = canonical_coverage_tier(ws.cell(row=target_row, column=cov_col).value if cov_col else None)
            
            if not cen_tier:
                cov_status = "not found on census"
            elif not inv_tier:
                cov_status = "not found on invoice"
            elif inv_tier == cen_tier:
                cov_status = "Matched"
            else:
                cov_status = "Mismatched"

            cell = ws.cell(row=target_row, column=disc_col)
            cell.value = f"Employee status : LLM Matched | Coverage status : {cov_status}"
            cell.fill = _FILL_LLM
            cell.font = _FONT
            cell.alignment = _CENTER

            match_count += 1
            # Keep track of successfully matched invoice names to delete their appended rows
            target_invoice_raw_names.add(target_invoice['raw_name'].strip().lower())
        else:
            logger.warning(f"Failed to apply match: {c_name} -> {i_name}. target_row={target_row}, target_invoice={bool(target_invoice)}, disc_col={disc_col}")

    # Delete the appended "Not on census" rows for the invoices we just matched
    rows_to_delete = set()
    if disc_col and target_invoice_raw_names:
        import re
        def clean_name_for_compare(name: str) -> str:
            s = name.strip().lower()
            if ',' in s:
                parts = [p.strip() for p in s.split(',')]
                s = f"{parts[1]} {parts[0]}" if len(parts) >= 2 else s
            s = re.sub(r"[^a-z0-9\s]", " ", s)
            s = re.sub(r"\s+", " ", s)
            return s.strip()

        def compact(name: str) -> str:
            """
            Normalize a name to a sorted, deduplicated, letter-only string.
            This handles two problems at once:
              1. Space splits: 'Uessugui Gomes' == 'Uessuguigomes' (same letters, different spaces)
              2. Doubled raw_name: Phase 3 stores raw_name as 'Rosanna Uessuguigomes Uessuguigomes, Rosanna'
                 (first+last+full concatenated). After dedup, both sides reduce to the same string.
            """
            s = name.strip().lower()
            if ',' in s:
                parts = [p.strip() for p in s.split(',')]
                s = f"{parts[1]} {parts[0]}" if len(parts) >= 2 else s
            tokens = re.sub(r"[^a-z\s]", " ", s).split()
            # Deduplicate while preserving first-occurrence order, then sort for order-independence
            seen = set()
            unique_tokens = []
            for t in tokens:
                if t not in seen:
                    seen.add(t)
                    unique_tokens.append(t)
            return "".join(sorted(unique_tokens))   # sorted → order-independent join

            
        target_invoice_cleaned = {clean_name_for_compare(n) for n in target_invoice_raw_names if n}
        target_invoice_compact  = {compact(n)               for n in target_invoice_raw_names if n}
        
        def get_name_from_row(row_idx: int) -> str:
            if name_col:
                val = ws.cell(row=row_idx, column=name_col).value
                return str(val).strip() if val else ""
            if first_col and last_col:
                f = str(ws.cell(row=row_idx, column=first_col).value or '').strip()
                l = str(ws.cell(row=row_idx, column=last_col).value  or '').strip()
                return f"{f} {l}".strip()
            # fallback to columns 2 and 3 if split, or column 1 if full
            f_val = str(ws.cell(row=row_idx, column=1).value or "").strip()
            s_val = str(ws.cell(row=row_idx, column=2).value or "").strip()
            t_val = str(ws.cell(row=row_idx, column=3).value or "").strip()
            if (not f_val or f_val.isdigit()) and s_val and t_val:
                return f"{s_val} {t_val}"
            return f_val

        for row_idx in range(1, ws.max_row + 1):
            disc_val = str(ws.cell(row=row_idx, column=disc_col).value or "").strip().lower()
            if "not on census" in disc_val:
                appended_name = get_name_from_row(row_idx)
                # Primary check: exact cleaned match
                # Secondary check: compact (space-stripped) match handles
                #   'Uessuguigomes' == 'Uessugui Gomes' and similar OCR/space issues
                if (clean_name_for_compare(appended_name) in target_invoice_cleaned
                        or compact(appended_name) in target_invoice_compact):
                    rows_to_delete.add(row_idx)
                    logger.info(
                        f"Marking row {row_idx} ('{appended_name}') for deletion "
                        f"— invoice name matched (space-agnostic)."
                    )

    for r in sorted(list(rows_to_delete), reverse=True):
        ws.delete_rows(r)

    if rows_to_delete:
        logger.info(f"Deleted {len(rows_to_delete)} appended 'Not on census' rows.")

    # --- NEW POST-PROCESSING: Clear plan/premium for 'WO' rows ---
    cleared_wo_count = 0
    if cov_col is not None and plan_col is not None and prem_col is not None:
        for r_idx in range(1, ws.max_row + 1):
            if r_idx in rows_to_delete: # Skip if row is already marked for deletion
                continue
            
            cov_val = str(ws.cell(row=r_idx, column=cov_col).value or '').strip().upper()
            if cov_val == 'WO':
                # Clear Plan Description
                ws.cell(row=r_idx, column=plan_col).value = None
                # Clear Monthly Total Premium
                ws.cell(row=r_idx, column=prem_col).value = None
                cleared_wo_count += 1
        if cleared_wo_count > 0:
            logger.info(f"Cleared plan and premium for {cleared_wo_count} 'WO' rows during post-processing.")
    # --- END NEW POST-PROCESSING ---

    if output_excel is None:
        output_excel = validated_excel.with_name(validated_excel.name.replace("VALIDATED_", "LLM_RESOLVED_"))
    
    wb.save(str(output_excel))
    logger.info(f"Saved LLM resolved Excel to {output_excel}")

    return {
        'status': 'completed',
        'matches': match_count,
        'output_path': str(output_excel)
    }

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Phase 4: LLM Fallback Resolver")
    parser.add_argument("validated_excel")
    parser.add_argument("audit_json")
    parser.add_argument("--output", default=None)
    parser.add_argument(
        "--template-type", dest="template_type", default="type3",
        choices=["type1", "type2", "type3"],
        help="Template type from flow_orchestrator. CH/SP skip guard only applies to type3 (engage)."
    )
    args = parser.parse_args()

    result = run_llm_resolution(
        Path(args.validated_excel),
        Path(args.audit_json),
        Path(args.output) if args.output else None,
        template_type=args.template_type,
    )
    if result.get('output_path'):
        print(f"LLM Output: {result['output_path']}")
