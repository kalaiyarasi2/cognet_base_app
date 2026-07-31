"""
RPVE_standalone.py
==================
Standalone FastAPI server for the RPVE Benefit Invoice Extractor.

RPVE = Resourcing · Prestige · Velocity · Engage

USAGE
-----
1. pip install -r requirements_RPVE.txt
2. Add OPENAI_API_KEY=sk-... to .env
3. python RPVE_standalone.py
4. Open http://localhost:8009

ENDPOINTS
---------
POST /extract          Upload PDF -> JSON + Excel download link
GET  /download/{file}  Download generated Excel
GET  /health           Health check
GET  /                 Serves RPVE_ui.html
"""

import asyncio
import logging
import logging.handlers
import sys
import os, re, json, shutil, uuid, time
from concurrent.futures import ThreadPoolExecutor
import pdfplumber
from contextlib import asynccontextmanager
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
from typing import List
from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

BASE_DIR   = Path(__file__).parent
UPLOAD_DIR = BASE_DIR / "rpve_uploads"
OUTPUT_DIR = BASE_DIR / "rpve_outputs"
JOBS_DIR   = BASE_DIR / "jobs"
UPLOAD_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
JOBS_DIR.mkdir(exist_ok=True)

# ══════════════════════════════════════════════════════════════════════════════
# CENTRALIZED SERVICE LOGGING
# ══════════════════════════════════════════════════════════════════════════════

def setup_logging():
    """
    Configure centralized logging so ALL backend output is captured in service.log.

    This captures:
      - All print() calls (via stdout/stderr redirect)
      - All Python logging (per-job loggers, uvicorn access/error)
      - Subprocess output logged via logger.debug()

    The log file rotates at 10 MB and keeps the last 5 backups.
    """
    log_path = BASE_DIR / "service.log"

    # ── Root logger — catches everything ──────────────────────────────────────
    root_logger = logging.getLogger()
    
    if root_logger.handlers:
        return  # Already configured

    root_logger.setLevel(logging.DEBUG)

    # File handler — rotating, 10 MB max, keep 5 backups
    file_handler = logging.handlers.RotatingFileHandler(
        str(log_path),
        maxBytes=10 * 1024 * 1024,  # 10 MB
        backupCount=5,
        encoding="utf-8",
    )
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(
        logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
    )
    root_logger.addHandler(file_handler)

    class SafeStreamWriter:
        """Wrapper to prevent UnicodeEncodeError in console output on Windows. [ignoring loop detection]"""
        def __init__(self, stream):
            self.stream = stream
        def write(self, data):
            try:
                self.stream.write(data)
            except UnicodeEncodeError:
                # Replace unsupported characters with ascii representation
                encoded = data.encode('ascii', errors='backslashreplace').decode('ascii')
                self.stream.write(encoded)
        def flush(self):
            try:
                self.stream.flush()
            except:
                pass

    # Console handler — keep normal terminal output too
    console_handler = logging.StreamHandler(SafeStreamWriter(sys.__stdout__))
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(
        logging.Formatter("%(message)s")
    )
    root_logger.addHandler(console_handler)

    # ── Redirect print() / stdout / stderr to the logger ─────────────────────
    import threading
    class _StreamToLogger:
        """File-like object that redirects writes to a Python logger with recursion protection."""
        _local = threading.local()

        def __init__(self, logger_instance, level):
            self._logger = logger_instance
            self._level = level

        def write(self, msg):
            if msg and msg.strip():
                # Detect recursion loop
                if getattr(self._local, 'lock', False):
                    sys.__stderr__.write(msg)
                    return
                
                try:
                    self._local.lock = True
                    for line in msg.rstrip("\n").splitlines():
                        self._logger.log(self._level, line)
                finally:
                    self._local.lock = False

        def flush(self):
            pass

        def isatty(self):
            return False

    sys.stdout = _StreamToLogger(logging.getLogger("rpve.stdout"), logging.INFO)
    sys.stderr = _StreamToLogger(logging.getLogger("rpve.stderr"), logging.WARNING)

    # ── Make per-job loggers propagate to root (so they appear in service.log) ─
    logging.getLogger("rpve").propagate = True

    # ── Reduce noise from noisy third-party libraries ─────────────────────────
    logging.getLogger("httpcore").setLevel(logging.WARNING)
    logging.getLogger("httpx").setLevel(logging.WARNING)
    logging.getLogger("openai").setLevel(logging.WARNING)
    logging.getLogger("PIL").setLevel(logging.WARNING)
    logging.getLogger("pdfminer").setLevel(logging.WARNING)

    logging.getLogger("rpve.startup").info(
        "Service logging initialised → %s", log_path
    )


# Initialise logging BEFORE anything else runs
setup_logging()

POPPLER_PATH = os.getenv("POPPLER_PATH")
if POPPLER_PATH and os.path.exists(POPPLER_PATH):
    if POPPLER_PATH not in os.environ["PATH"]:
        os.environ["PATH"] += os.pathsep + POPPLER_PATH
    print(f"[RPVE] Poppler path added to PATH: {POPPLER_PATH}")

client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# ══════════════════════════════════════════════════════════════════════════════
# UNIFIED FIELD SCHEMA
# ══════════════════════════════════════════════════════════════════════════════

# A single, unified schema for all extractions as per user requirements.
# The specialized "engage" prompt is also designed to return these fields.
UNIFIED_FIELDS = [
    "full_name",
    "first_name",
    "middal_name",
    "last_name",
    "coverage",
    "plan_name",
    "plan_type",
    "current_premium",
    "adjustment_amount",
    "birth_date",
    "gender",
    "home_zip_code",
    "billing_period",
]

# The employee fields dictionary now only distinguishes between ADP and generic.
EMPLOYEE_FIELDS = {
    "engage":       UNIFIED_FIELDS,
    "generic":      UNIFIED_FIELDS,
    "datalink_emi": UNIFIED_FIELDS,   # Data Link EMI uses the same unified schema
}

# Simplified summary fields. The generic prompt will attempt to find these.
SUMMARY_FIELDS = {
    "engage":  ["COMPANY_NAME", "INVOICE_NUMBER", "BILLING_DATE", "DUE_DATE", "REFERENCE_NUMBER"],
    "generic": ["COMPANY_NAME", "INVOICE_NUMBER", "BILLING_DATE", "DUE_DATE"],
}

SUB_TYPE_LABELS = {
    "engage":  "Engage (ADP TotalSource)",
    "generic": "Generic Document",
}

HEADER_COLOURS = {
    "engage":  "3C3489",
    "generic": "666666",
}

# ══════════════════════════════════════════════════════════════════════════════
# KEYWORD CLASSIFIER
# ══════════════════════════════════════════════════════════════════════════════

# Simplified to identify ADP ("engage") and ACSA documents.
KEYWORDS = {
    "engage": ["TOTALSOURCE", "TOTALSOURCE BENEFITS INVOICE", "TOTALSOURCE® BENEFITS INVOICE", "NCT3-EPO", "ADP", "ADP, INC", 
               "ASSOCIATION OF COMMUNITY SERVICE", "ACSA", "ACSA GROUP INSURANCE", "HEALTHNET", "HEALTH NET"],
    # Data Link EMI carrier
    "datalink_emi": ["DATA LINK EMI", "DATALINK EMI", "DATALINKEMI"],
}

# ══════════════════════════════════════════════════════════════════════════════
# LLM PROMPTS - output keys match EMPLOYEE_FIELDS exactly
# ══════════════════════════════════════════════════════════════════════════════

PROMPTS = {

    "deduction_roster": """
You are extracting data from a BENEFIT DEDUCTION ROSTER.

Extract a SUMMARY and EMPLOYEES array.

SUMMARY: company_name, invoice_number, billing_date, total_amount_due

EMPLOYEE RECORDS:
For each employee plan line, you MUST extract:
  first_name: member first name
  last_name: member last name  
  coverage: coverage level (e.g. EE, ES, FAM, EC)
  plan_name: the "Plan Description" column value
  current_premium: YOU MUST TAKE THE VALUE FROM THE "Monthly Premium Total" COLUMN.
  
🔹 CRITICAL RULE:
There are two "Total" columns: "Monthly Premium Total" and "Pay Period Amount Total".
YOU MUST USE THE "Monthly Premium Total".
DO NOT USE THE "Pay Period Amount Total".

Example:
If a row shows "525.00" under Monthly Premium Total and "262.50" under Pay Period Amount Total, current_premium MUST BE "525.00".

Use "" for missing values. Return ONLY valid JSON.

{
  "summary": {"company_name":"","invoice_number":"","billing_date":"","total_amount_due":""},
  "employees": [{"first_name":"","last_name":"","coverage":"","plan_name":"","current_premium":""}]
}

PDF TEXT: {text}
""",

    "engage": """
You are extracting data from a group health insurance invoice (ADP TotalSource, ACSA, or similar).

Extract a SUMMARY and EMPLOYEES array.

SUMMARY: company_name, invoice_number, billing_date, due_date, reference_number, total_amount_due

FOR SUMMARY-ONLY INVOICES (like ACSA):
If the invoice shows only summary financial data (no individual employee roster), create ONE consolidated record:
  first_name: "SUMMARY"
  last_name: "TOTAL" 
  coverage: "EMPLOYER"
  plan_name: "HEALTH PLAN"
  coverage_option: billing period (e.g. "03/01/2026 through 03/31/2026")
  current_premium: total amount due

FOR EMPLOYEE ROSTER INVOICES (like ADP TotalSource):
CRITICAL: For EACH employee, you MUST extract EVERY plan line they have — do NOT skip any lines.
Each plan line for each employee becomes a separate record. Include ALL of:
  - Medical / Health plans (HMO, PPO, HSA, EPO, HDHP, KAI-*, BC-*, ANT*, KAIPER, ANTHMO, ANTPPO, etc.)
  - Dental plans (AETDEN, DEN, PPO Dental, etc.)
  - Vision plans (VSP, Vision, etc.)
  - Life Insurance plans (LIF, METLDI LIF, Basic Life, etc.)
  - LTD / Disability plans (LTD, METLDI LTD, etc.)
  - Any other benefit plan line with a dollar amount

Fields per record:
  first_name: member first name
  last_name: member last name  
  coverage: EXACT coverage tier/code (e.g. Employee, Family, EE+1, E, ES, ESC, EC, E1D, etc.)
  plan_name: insurance category/type — one of: Medical, Dental, Vision, Life, LTD, Other
  coverage_option: specific insurance product name (e.g. "BC-HMO 30-100-SCA", "KAI-HMO 30-South-CA", "VSP- Choice Vision Plan")
  current_premium: dollar amount for that INDIVIDUAL plan line (NOT the Total row)

EXTRACTION RULES:
1. COMPLETENESS IS MANDATORY: Every employee MUST have all their plan lines extracted. If an employee has Medical + Dental + Vision + Life + LTD, that is 5 separate records.
2. NEVER SKIP MEDICAL LINES: If you see plan codes like ANTHMO, ANTPPO, ANTBCB, KAIPER, KAIHMO, BCPPO, BCHMO or any HMO/PPO/HSA plan, that is a Medical plan — ALWAYS include it.
3. DO NOT INCLUDE TOTAL ROWS: Rows labeled "Total" with a sum of all plans are NOT individual plan lines — skip them.
4. EE ID ROWS: Lines starting with "EE ID:" are identifiers for the employee, not a separate plan. They belong to the employee whose name appeared just before.
5. DETECT FORMAT: Look for individual employee names vs summary-only financial data.
6. FINANCIAL DATA: Prior balance, adjustments, current cost, admin fees, total due.
7. DATES: Billing date, due date, billing period.
8. BENEFIT DEDUCTION ROSTER: If the document is titled "BENEFIT DEDUCTION ROSTER", you MUST extract the value from the "Monthly Premium" -> "Total" column as the `current_premium`. STRICTLY FORBIDDEN: Do NOT use the "Pay Period Amount" values for the premium.

Use "" for missing values. Return ONLY valid JSON.

{
  "summary": {"company_name":"","invoice_number":"","billing_date":"","due_date":"","reference_number":"","total_amount_due":""},
  "employees": [{"first_name":"","last_name":"","coverage":"","plan_name":"","coverage_option":"","current_premium":""}]
}

PDF TEXT: {text}
""",

}

# ══════════════════════════════════════════════════════════════════════════════
# FALLBACK EXTRACTION FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def extract_engage_fallback(text: str) -> dict:
    """
    Fallback regex-based extractor for engage subtype when LLM extraction fails.
    Handles both ADP roster format and ACSA summary format.
    """
    import re
    
    # Initialize result structure
    result = {
        "summary": {
            "company_name": "",
            "invoice_number": "",
            "billing_date": "",
            "due_date": "",
            "reference_number": "",
            "total_amount_due": ""
        },
        "employees": []
    }
    
    try:
        text_upper = text.upper()
        
        # Check if this is ACSA format
        if "ASSOCIATION OF COMMUNITY SERVICE" in text_upper or "ACSA" in text_upper:
            return extract_acsa_summary_fallback(text)
        
        # For ADP/TotalSource format - extract basic summary info
        # Extract company name (look for common patterns)
        company_patterns = [
            r'Company\s*:\s*(.+?)(?:\n|\r)',
            r'Employer\s*Name\s*:\s*(.+?)(?:\n|\r)',
            r'Client\s*:\s*(.+?)(?:\n|\r)'
        ]
        
        for pattern in company_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                result["summary"]["company_name"] = match.group(1).strip()
                break
        
        # Extract billing date
        date_match = re.search(r'Billing\s*Date\s*:\s*(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
        if date_match:
            result["summary"]["billing_date"] = date_match.group(1)
        
        # Extract total amount (look for final totals)
        total_patterns = [
            r'Total\s*Amount\s*Due\s*[:\s]*(\d{1,3}(?:,\d{3})*\.\d{2})',
            r'Grand\s*Total\s*[:\s]*(\d{1,3}(?:,\d{3})*\.\d{2})',
            r'Amount\s*Due\s*[:\s]*(\d{1,3}(?:,\d{3})*\.\d{2})'
        ]
        
        for pattern in total_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                result["summary"]["total_amount_due"] = match.group(1)
                break
        
        print(f"[RPVE] Engage fallback extraction: Found {len(result['employees'])} employee records")
        return result
        
    except Exception as e:
        print(f"[RPVE] Engage fallback extraction failed: {e}")
        return {"summary": {}, "employees": []}


def extract_acsa_summary_fallback(text: str) -> dict:
    """
    Fallback regex-based extractor for ACSA summary format when LLM extraction fails.
    """
    import re
    
    result = {
        "summary": {
            "company_name": "",
            "invoice_number": "",
            "billing_date": "",
            "due_date": "",
            "reference_number": "",
            "total_amount_due": ""
        },
        "employees": []
    }
    
    try:
        # Extract company name for ACSA
        company_match = re.search(r'Association\s+of\s+Community\s+Service\s+Agencies?[\s\-]*(.+?)(?:\n|\r)', text, re.IGNORECASE)
        if company_match:
            result["summary"]["company_name"] = company_match.group(1).strip()
        
        # Extract billing period
        period_match = re.search(r'Billing\s+Period\s*[:\-\s]*(\d{2}/\d{2}/\d{4})[\s\-]*(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
        if period_match:
            result["summary"]["billing_date"] = period_match.group(1)
            result["summary"]["due_date"] = period_match.group(2)
        
        # Extract total amount
        total_match = re.search(r'Total\s+Premium\s*[:\-\s]*\$(\d{1,3}(?:,\d{3})*\.\d{2})', text, re.IGNORECASE)
        if total_match:
            result["summary"]["total_amount_due"] = total_match.group(1)
        
        print(f"[RPVE] ACSA fallback extraction: Found {len(result['employees'])} employee records")
        return result
        
    except Exception as e:
        print(f"[RPVE] ACSA fallback extraction failed: {e}")
        return {"summary": {}, "employees": []}


# ══════════════════════════════════════════════════════════════════════════════
# CORE FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════════

def extract_text(pdf_path: Path, max_pages: int = 1000) -> str:
    """
    Robustly extracts text from a PDF by iterating through pages and using 
    multiple engines (pdfplumber, fitz, and OCR) to ensure full capture.
    Handles rotated and reversed text mapping issues.
    """
    text = ""
    print(f"[RPVE] Extracting text from {pdf_path.name}...")

    # Keywords we EXPECT to find in a valid RPVE document page
    VALID_KEYWORDS = [
        "TOTALSOURCE", "PAYCHEX", "AETNA", "KAISER", "UNITEDHEALTHCARE", "INVOICE", "BILLING", 
        "PREMIUM", "AMOUNT DUE", "PAGE", "EMPLOYEE", "MEMBERS", "CURRENT DETAIL", 
        "RETRO DETAIL", "ADJUSTMENT DETAIL", "MEDICA", "ADP", "BLUE CROSS", "CIGNA", "GUARDIAN"
    ]

    try:
        import pdfplumber
        import fitz
        import pytesseract
        from PIL import Image
        
        with pdfplumber.open(str(pdf_path)) as pdf:
            with fitz.open(pdf_path) as doc:
                pages_to_extract = min(max_pages, len(doc))
                for i in range(pages_to_extract):
                    page_text = ""
                    
                    # 1. Try pdfplumber (best for layout preservation)
                    try:
                        if i < len(pdf.pages):
                            plumber_page = pdf.pages[i]
                            p_text = plumber_page.extract_text(layout=True) or ""
                            if len(p_text.strip()) > 100 and any(kw in p_text.upper() for kw in VALID_KEYWORDS):
                                page_text = p_text
                    except Exception as e:
                        print(f"  [PAGE {i+1}] pdfplumber error: {e}")

                    # 2. Fallback to fitz (PyMuPDF) if pdfplumber is empty or fails keywords
                    if not page_text.strip():
                        try:
                            fitz_page = doc[i]
                            f_text = fitz_page.get_text() or ""
                            if len(f_text.strip()) > 50 and any(kw in f_text.upper() for kw in VALID_KEYWORDS):
                                page_text = f_text
                        except Exception as e:
                            print(f"  [PAGE {i+1}] fitz failed: {e}")

                    # 3. Last Resort: Robust High-Accuracy OCR (handles scanned/rotated text)
                    if not page_text.strip():
                        print(f"  [PAGE {i+1}] Running Robust OCR Fallback...")
                        try:
                            # Render at 3x zoom (216 dpi) for high accuracy
                            pix = doc[i].get_pixmap(matrix=fitz.Matrix(3, 3))
                            img = Image.frombytes('RGB', [pix.width, pix.height], pix.samples)
                            
                            # A. Try Standard OSD detection first
                            try:
                                osd = pytesseract.image_to_osd(img)
                                rotation = re.search(r'Rotate: (\d+)', osd)
                                if rotation:
                                    angle = int(rotation.group(1))
                                    if angle != 0:
                                        print(f"    [OSD] Correcting {angle}° rotation...")
                                        img = img.rotate(-angle, expand=True)
                            except:
                                pass

                            # B. Try OCR at current orientation
                            page_text = pytesseract.image_to_string(img, config='--psm 6')

                            # C. Brute-Force Rotation Fallback (if keywords missing)
                            if not any(kw in page_text.upper() for kw in VALID_KEYWORDS):
                                print(f"    [PAGE {i+1}] Keywords not found. Retrying 90/180/270 degree rotations...")
                                ori_img = img.copy()
                                for rot in [90, 180, 270]:
                                    img_rot = ori_img.rotate(rot, expand=True)
                                    test_text = pytesseract.image_to_string(img_rot, config='--psm 6')
                                    if any(kw in test_text.upper() for kw in VALID_KEYWORDS):
                                        page_text = test_text
                                        print(f"    [PAGE {i+1}] Success at {rot}° rotation.")
                                        break
                            
                            # D. Final check: if still empty, use psm 3 (standard)
                            if not page_text.strip():
                                page_text = pytesseract.image_to_string(img, config='--psm 3')

                        except Exception as e:
                            print(f"  [PAGE {i+1}] OCR failed: {e}")

                    text += f"--- Page {i + 1} ---\n" + page_text + "\n"
    except Exception as e:
        print(f"[RPVE] Global extraction error: {e}")

    extracted_upper = text.upper()
    text_quality_score = assess_text_quality(text)
    
    # ── CARRIER-SPECIFIC EXTRACTION UPGRADE ───────────────────────────
    # If this is a wide-layout carrier like BlueCross or IBX, standard 
    # vertical extraction often shreds rows. We force the high-accuracy 
    # layout-preserving OCR fallback for these specifically.
    WIDE_KEYWORDS = ["BLUE CROSS", "INDEPENDENCE", "BCBS", "CAPITAL BLUE", "IBX", "BLUE SHIELD", "UNITEDHEALTHCARE", "UHC", "UNITED HEALTHCARE"]
    is_wide_carrier = any(kw in extracted_upper for kw in WIDE_KEYWORDS)

    print(f"[RPVE] Extraction check: quality={text_quality_score:.2f}, is_wide={is_wide_carrier}")

    if (not text.strip() 
        or not any(kw in extracted_upper for kw in VALID_KEYWORDS)
        or text_quality_score < 0.75
        or is_wide_carrier):
        
        reason = ""
        if is_wide_carrier:
            reason = "wide carrier layout detected (forcing high-accuracy pass)"
        elif not text.strip():
            reason = "text is empty"
        elif not any(kw in extracted_upper for kw in VALID_KEYWORDS):
            reason = f"keywords not found (quality: {text_quality_score:.2f})"
        else:
            reason = f"text quality too low ({text_quality_score:.2f} < 0.75 threshold)"
        
        print(f"[RPVE] Triggering high-accuracy fallback: {reason}")
        rostaing_text = extract_text_with_rostaing(pdf_path)
        
        # ── INTELLIGENT FALLBACK VALIDATION ───────────────────────────
        # Only use Rostaing text if it's actually "better" or at least 
        # comparable in volume. If Rostaing fails on rotated pages 
        # (producing less text than standard), we stick to standard.
        if rostaing_text and rostaing_text.strip():
            # Lower threshold: OCR text is often much cleaner/shorter than "raw" PDF text
            if len(rostaing_text) >= (len(text) * 0.20):
                print(f"[RPVE] Using Rostaing OCR result ({len(rostaing_text)} chars).")
                return rostaing_text
            else:
                print(f"[RPVE] Rostaing result suspiciously short ({len(rostaing_text)} vs {len(text)}). Sticking to standard extraction.")

    return text


def assess_text_quality(text: str) -> float:
    """
    Score text quality (0.0 to 1.0) to detect OCR corruption.
    Low score = high corruption (garbled chars, broken tables, fragmentation).
    """
    if not text or len(text) < 50:
        return 0.0
    
    lines = text.split('\n')
    line_count = len(lines)
    
    # 1. Check for garbled/non-ASCII characters (corruption indicator)
    garbled_count = 0
    for char in text:
        if ord(char) > 127 and char not in 'àáâãäåèéêëìíîïòóôõöùúûüýÿªºñ—–':
            garbled_count += 1
    
    garbled_ratio = garbled_count / len(text) if len(text) > 0 else 0
    
    # 2. Check for repeated fragmented lines (broken table markers)
    fragment_pattern = r'\s*\|\s*|\s+\[.{1,3}\]\s+'
    fragment_count = len(re.findall(fragment_pattern, text))
    fragment_ratio = fragment_count / max(line_count, 1)
    
    # 3. Check for very short lines (fragmentation sign)
    short_lines = sum(1 for line in lines if len(line.strip()) < 5 and line.strip())
    short_line_ratio = short_lines / max(line_count, 1) if line_count > 0 else 0
    
    # 4. Check for repeated consecutive lines (OCR duplication artifact)
    duplicate_lines = 0
    for i in range(1, len(lines)):
        if lines[i].strip() and lines[i-1].strip() and lines[i].strip() == lines[i-1].strip():
            duplicate_lines += 1
    
    duplicate_ratio = duplicate_lines / max(line_count - 1, 1)
    
    # Weighted quality score
    quality = 1.0
    quality -= min(0.3, garbled_ratio * 3)        # Up to 30% penalty for non-ASCII
    quality -= min(0.4, fragment_ratio * 3)       # Up to 40% penalty for fragments (increased weight)
    quality -= min(0.25, short_line_ratio * 2)    # Up to 25% penalty for short lines
    quality -= min(0.25, duplicate_ratio * 2)     # Up to 25% penalty for duplicates
    
    quality = max(0.0, min(1.0, quality))
    print(f"[RPVE] Text quality assessment: {quality:.2f} (garbled: {garbled_ratio:.2%}, fragments: {fragment_ratio:.2%}, short_lines: {short_line_ratio:.2%}, duplicates: {duplicate_ratio:.2%})")
    
    return quality


def extract_text_with_rostaing(pdf_path: Path) -> str:
    """Fallback PDF text extraction using rostaing-ocr when the standard path is noisy."""
    if pdf_path.suffix.lower() != '.pdf':
        print(f"[RPVE] Skipping rostaing-ocr fallback: {pdf_path.name} is not a PDF file.")
        return ""

    try:
        from schema_ocr import SchemaOCRExtractor
    except Exception as e:
        print(f"[RPVE] Could not import schema_ocr for rostaing fallback: {e}")
        return ""

    if shutil.which("tesseract") is None:
        print("[RPVE] Tesseract not found in PATH. rostaing-ocr may still work, but OCR accuracy could be reduced.")

    try:
        extractor = SchemaOCRExtractor(pdf_path)
        text = extractor.extract_layout_text(save_debug_output=True)
        if text and text.strip():
            print(f"[RPVE] Rostaing OCR fallback produced {len(text.splitlines())} lines of text.")
            return text
        print("[RPVE] Rostaing OCR fallback returned empty text.")
    except Exception as e:
        print(f"[RPVE] Rostaing OCR fallback failed: {e}")

    return ""


def clean_invoice_text(text: str) -> str:
    """
    Remove common noise like page footers, headers, and copyright notices.
    Also handles 'orphaned' Total rows that can confuse the LLM at page boundaries.
    """
    if not text: return ""
    lines = text.split('\n')
    cleaned_lines = []
    
    # Regex to detect page footers, headers, copyright notices, and orphaned Totals
    # We remove "Total" rows that have NO provider/plan data if they appear near headers
    header_footer_pattern = re.compile(r"""
        ^\s*page\s+\d+\s+of\s+\d+\s*$|
        copyright\s+©\s+.*adp,\s+inc|
        ^\s*Name\s+Provider\s+Plan\s+Coverage\s+Type\s+Month\s+Cost\s*$|
        ^\s*Total\s+\$[\d,]+\.\d{2}\s*$
    """, re.IGNORECASE | re.VERBOSE)

    for line in lines:
        # If the line doesn't match the noise pattern, keep it
        if not header_footer_pattern.search(line):
            cleaned_lines.append(line)
            
    return '\n'.join(cleaned_lines)

def group_indented_lines(text: str) -> str:
    """
    DYNAMICAL SOLUTION: 
    Detects lines starting with significant whitespace OR 'EE ID:' and 
    stitches them to the preceding 'Name' row. 
    """
    if not text: return ""
    lines = text.split('\n')
    output = []
    current_block = ""

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
            
        # DYNAMICAL RULES:
        # 1. Any line starting with "EE ID:" belongs to the employee above.
        # 2. Any line starting with 10+ spaces (indentation > name) belongs to employee above.
        # 3. Any line starting with "Total" (at sub-indentation) belongs to employee above.
        
        is_sub_row = (
            stripped.startswith("EE ID:") or 
            line.startswith("          ") or 
            (line.startswith("     ") and stripped.startswith("Total"))
        )
        
        if is_sub_row and current_block:
            current_block += " [SUB-ROW] " + stripped
        else:
            if current_block:
                output.append(current_block)
            current_block = line
            
    if current_block:
        output.append(current_block)
        
    return '\n'.join(output)


def classify(text: str) -> str:
    """Classifies the document for extraction type."""
    t = text.upper()
    
    # Priority 0: Check for Benefit Deduction Roster
    if "BENEFIT DEDUCTION ROSTER" in t:
        print("[RPVE] Detected Benefit Deduction Roster format.")
        return "deduction_roster"

    # Priority 1: Check for ADP/TotalSource (original engage)
    if "TOTALSOURCE" in t or "ADP" in t:
        print("[RPVE] Detected ADP TotalSource format.")
        return "engage"
    
    # Priority 2: Check for ACSA/Community Service Agencies (treat as engage)
    if "ASSOCIATION OF COMMUNITY SERVICE" in t or "ACSA" in t:
        print("[RPVE] Detected ACSA Health Plan format, treating as engage type.")
        return "engage"
    
    # Priority 3: Check for other specific formats
    for sub_type, kwlist in KEYWORDS.items():
        if any(kw in t for kw in kwlist):
            print(f"[RPVE] Classified as {sub_type.upper()} format.")
            return sub_type
    
    # Check for summary-only invoice characteristics (treat as engage)
    if "EMPLOYER NAME" in t and "BILLING DATE" in t and "TOTAL AMOUNT" in t:
        print("[RPVE] Detected summary-only invoice format, treating as engage type.")
        return "engage"

    # Priority 4: Check for UnitedHealthcare
    if "UNITEDHEALTHCARE" in t or "UHC" in t:
        print("[RPVE] Detected UnitedHealthcare format.")
        return "generic"
    
    print("[RPVE] No specific keywords matched. Using GENERIC extractor.")
    return "generic"


def clean_invoice_text(text: str) -> str:
    """
    Cleans the extracted invoice text by removing headers, footers, and other noise
    that can disrupt the LLM's parsing of continuous employee data.
    """
    lines = text.split('\n')
    cleaned_lines = []
    
    # Regex to detect page footers, headers, and copyright notices
    header_footer_pattern = re.compile(r"""
        ^\s*page\s+\d+\s+of\s+\d+\s*$|
        copyright\s+©\s+.*adp,\s+inc|
        ^\s*Name\s+Provider\s+Plan\s+Coverage\s+Type\s+Month\s+Cost\s*$|
        ^\s*Consolidated\s+Customer\s+No.*$|
        ^\s*Customer\s+No.*Invoice\s+No.*$|
        ^\s*Invoice\s+Date:.*$|
        ^\s*Bill\s+Group:.*$|
        ^\s*Coverage\s+Period:.*$|
        ^\s*Due\s+Date:.*$|
        ^\s*Details\s*$|
        ^\s*Current\s+Detail\s+-.*$|
        ^\s*Policy\s+Name\s+Plan\s+ID.*$|
        ^\s*No\.\s+\(000's\)\s+Amount.*$|
        ^\s*Coverage\s+Type\s+Status\s+Code\s*$|
        ^\s*E\s+Employee\s+Only.*$|
        ^\s*ES\s+Employee\s+and\s+Spouse.*$|
        ^\s*ESC\s+Employee\s+and\s+Family.*$|
        ^\s*EC\s+Employee\s+and\s+Child.*$|
        ^\s*E1D\s+Employee\s+and\s+One.*$|
        ^\s*E2D\s+Employee\s+and\s+Two.*$|
        ^\s*E3D\s+Employee\s+and\s+Three.*$|
        ^\s*Questions\?.*$|
        ^\s*d50\s+WWA.*$|
        ^\s*w\.\s+n\s+asennces.*$
    """, re.IGNORECASE | re.VERBOSE)

    for line in lines:
        # If the line doesn't match the pattern, keep it
        if not header_footer_pattern.search(line):
            cleaned_lines.append(line)
            
    # Rejoin the lines
    cleaned_text = '\n'.join(cleaned_lines)
    
    return cleaned_text


def split_multi_invoice_text(text: str) -> list[str]:
    # Find all occurrences of page 1 of X
    pattern = re.compile(r'(page\s+1\s+of\s+\d+)', re.IGNORECASE)
    matches = list(pattern.finditer(text))
    if not matches:
        return [text]
        
    parts = []
    for i in range(len(matches)):
        start = matches[i].start()
        end = matches[i+1].start() if i + 1 < len(matches) else len(text)
        parts.append(text[start:end])
    return parts


def split_text_by_pages(text: str) -> list[str]:
    """Splits text using '--- Page \d+ ---' markers."""
    pattern = re.compile(r'(--- Page \d+ ---)', re.IGNORECASE)
    matches = list(pattern.finditer(text))
    if not matches:
        return [text]
    
    pages = []
    for i in range(len(matches)):
        start = matches[i].start()
        end = matches[i+1].start() if i+1 < len(matches) else len(text)
        pages.append(text[start:end])
    return pages


def extract_with_llm(sub_type: str, text: str, ev_mode: bool = False) -> dict:
    """
    Calls the LLM to extract structured summary and employee data.
    Uses carrier-specific prompts if available, otherwise falls back to a standard prompt.
    """               
    # Check for multiple concatenated invoices
    parts = split_multi_invoice_text(text)
    if len(parts) > 1:
        print(f"[RPVE] Found {len(parts)} multi-invoice parts. Processing each part separately...")
        all_employees = []
        final_summary = {}
        for idx, part in enumerate(parts, 1):
            print(f"[RPVE] Processing invoice part {idx}/{len(parts)}...")
            part_result = extract_with_llm(sub_type, part, ev_mode)
            part_summary = part_result.get("summary", {})
            part_employees = part_result.get("employees", [])
            for emp in part_employees:
                # Copy summary fields to the employee dictionary
                for k, v in part_summary.items():
                    emp[k.lower()] = v
            all_employees.extend(part_employees)
            if not final_summary:
                final_summary = part_summary
            else:
                # Accumulate/Merge total_amount_due if applicable
                t_keys = ["total_amount_due", "total_cost", "grand_total"]
                for tk in t_keys:
                    if tk in part_summary and tk in final_summary:
                        try:
                            val1 = float(re.sub(r'[^\d.-]', '', str(final_summary[tk])))
                            val2 = float(re.sub(r'[^\d.-]', '', str(part_summary[tk])))
                            final_summary[tk] = f"${(val1 + val2):,.2f}"
                        except:
                            pass
        return {
            "summary": final_summary,
            "employees": all_employees
        }

    # Clean the text to handle multi-page table fragmentation
    text = clean_invoice_text(text)
    
    # DYNAMICAL STITCHING: Group indented rows before sending to LLM
    if sub_type in ["engage", "prestige", "velocity"]:
        text = group_indented_lines(text)

    # 1. Determine which prompt to use
    prompt_template = PROMPTS.get(sub_type)
    
    if not prompt_template:
        # Fallback to the Standard / Generic Prompt
        prompt_template = """
You are a data extraction engine processing a group insurance invoice.

🔹 CAPTURE ALL MEMBERS & ADJUSTMENTS (CRITICAL)
This invoice may list members in a ROSTER format where each row is a separate individual (Subscriber, Spouse, Dependent).
You MUST extract EVERY person listed in ANY section:
  - CURRENT DETAIL section
  - RETRO DETAIL section
  - ADJUSTMENT DETAIL section
  - Any other detail section in the document

🔹 NEGATIVE VALUES (CRITICAL):
  - If a value is negative (e.g. $-100.00), you MUST preserve the minus sign in the current_premium or adjustment_amount field.

🔹 ADP FORMAT SPECIFIC RULES (APPLY ONLY IF "TOTALSOURCE", "ADP", OR "NCT3-EPO" IS PRESENT)
If the document is EXPLICITLY identified as an ADP invoice (e.g. ADP TotalSource format), you MUST apply these strict rules. If it is NOT an ADP file, ignore these specific constraints and extract EVERY record regardless of amount:

1. Plan Name Extraction (CRITICAL for ADP):
Extract ONLY the exact, valid ADP plan name.
Do NOT extract random text near plan sections, headers, footers, or unrelated labels.
✅ Plan name must belong to a defined benefits section, be consistent across employee entries, and appear as a clear plan title.

🔹 DO NOT ABBREVIATE OR TRUNCATE PLAN NAMES (CRITICAL):
The plan name MUST match the FULL string found in the "Plan" column of the PDF.

🔹 ADP LAYOUT STRUCTURE:
Data is typically organized as a Header Row (Name + First Plan) followed by Indented Rows (Additional Plans). 
The pre-processor has marked these as '[SUB-ROW]'. You MUST split these [SUB-ROW] markers into separate individual plan records for the same employee.

🔹 NAME FORMATTING (CRITICAL):
Names are often printed as "LastName, FirstName" or "LastName, FirstName Middle" (e.g. "Smith, John Adam"). Properly identify and split the `last_name` and `first_name` without inverting them.

🔹 OUTPUT FIELDS (13 fields per person)
- full_name
- first_name
- middal_name (Middle Name)
- last_name
- coverage (e.g. ES / EC / FAM / EE / SP / CH)
- plan_name (FULL plan/product description — do NOT truncate). IMPORTANT: Plan names often wrap across multiple lines. You MUST concatenate these into a single string (e.g., "Non-Contributory 25K Flat Basic Life EE Only"). CRITICAL: Terms like "SINGLE", "FAMILY", "P&C", "PARENT & CHILD", "H/W" are COVERAGE TIERS, not plan names. NEVER put these in plan_name. CRITICAL: Do NOT infer or hallucinate a plan_name from column headers (e.g. if the only column is "Medical", do NOT set plan_name to "Medical"). If no explicit plan name is printed per-row in the document, leave plan_name as null.
- plan_type (insurance category: Medical, Dental, Vision, etc.)
- current_premium: The individual plan line cost for that specific plan row. (**CRITICAL FOR UHC/UHC NA**: Use the "Totals -> Total" column for single-line rows, but use "Charge Amount" if multiple distinct plan rows exist for the same member.)
- adjustment_amount: Any adjustment amount listed.
- birth_date
- gender (M / F — infer if not present)
- home_zip_code
- billing_period: The start and end date of the billing cycle for the line item (e.g., "01/01/2024 - 01/31/2024").

🔹 MAXIMUM RECALL (CRITICAL):
- **Completeness is the HIGHEST priority.** You MUST extract EVERY individual enrollment row found in the document.
- Do NOT skip any rows. Even if a row has partial data, extract what is available.
- If the document contains a roster (like IBX or BCBS), expect dozens of members. You must continue until the very end of the list.

🔹 PAGE BREAK CONTINUATIONS (CRITICAL):
If you see an employee's data interrupted or continued across pages, associate all subsequent plan lines with that specific employee until a new Name is encountered. Do NOT create unnamed records for orphaned plan lines.

🔹 KEY RULES:
- One row per individual member.
- Strictly adhere to JSON format.
- Do not hallucinate, but do not omit valid rows.

🔹 UNITEDHEALTHCARE / GENERIC INVOICE PLAN NAMES (CRITICAL RULE):
In UnitedHealthcare and similar invoices, the Plan Name column is frequently split across two lines:
  - Line 1 contains the alphanumeric plan prefix/code (e.g., "P15003050I8021B_6700", "HP6000I7025B_670079", or "P2503050I8022B_67006").
  - Line 2 contains the plan suffix and description (e.g., "70 - Max Claims Liability", "3 - Max Claims Liability", or "Max Claims Liability").
You MUST NEVER treat the alphanumeric prefix as an ignored group/policy code. You MUST concatenate Line 1 and Line 2 together into a single complete string (e.g., "P15003050I8021B_670070 - Max Claims Liability" or "P2503050I8022B_670063 - Max Claims Liability"). Do NOT truncate or omit the alphanumeric prefix!

🔹 INSPERITY / MANIFEST MEDEX:
- If column headers include "Coverage Type" and "Coverage Option", map "Coverage Type" -> `plan_type` and "Coverage Option" -> `plan_name`. Do not mix them up.

🔹 WARWICK / DEDUCTION REGISTER:
- If column headers include "Ded Code" or "Benefit Plan", map "Benefit Plan" to `plan_name` and "Ded Code" to `plan_type`.

🔹 KARPEN_STEEL_PRODUCTS:
- For this carrier, you MUST map the value from the "Total Premium" column to `adjustment_amount`.

🔹"DATA LINK EMI", "DATALINK EMI", "DATALINKEMI" :
- **CRITICAL: EXTRACT ONLY THE "Medical" COLUMN VALUE. ALL OTHER COLUMNS ARE FORBIDDEN.**
- FORBIDDEN columns (never extract these): "Total Due", "Dental", "Vision", "Garner HRA". Ignore them completely.
- For REGULAR rows (current billing):
    - `current_premium` = value from "Medical" column ONLY (e.g. $334.78, $0.00).
    - `adjustment_amount` = null.
- For RETRO ACTIVE ADJUSTMENT rows (negative values / retroactive section):
    - `current_premium` = null.
    - `adjustment_amount` = value from "Medical" column ONLY (e.g. $-334.78).
- If a member has Medical=$0.00, then current_premium MUST be "$0.00" — do NOT substitute Dental or Vision or Total Due.
- Example: Medical=$0.00, Dental=$58.10, Vision=$0.00, Total Due=$58.10 → current_premium="$0.00" ✅ NOT "$58.10" ❌
- Example: Medical=$334.78, Dental=$27.90, Vision=$7.70, Total Due=$412.38 → current_premium="$334.78" ✅ NOT "$412.38" ❌
- **If you return Total Due, Dental, Vision, or Garner HRA in any field, you have failed the task.**

🔹 CIGNA "BILLING DETAIL BY GROUP" SPECIAL RULES (CRITICAL):
- If the document header contains "BILLING DETAIL BY GROUP" and/or "cigna healthcare", you MUST extract the value from the "Total (4)" column (the far-right total column) as the `current_premium`.
- **STRICTLY FORBIDDEN:** Do NOT use the "Medical" column value for `current_premium` on these forms. The "Medical" column only shows a partial amount.
- The "Total (4)" column = Medical + Claims Funding combined. This is the correct premium to extract.
- Example: If Medical=$275.67, Amount Due=$275.67, Claims Funding=$473.35, Total (4)=$749.02 → current_premium="$749.02" ✅ NOT "$275.67" ❌
- Example: If Medical=$863.08, Amount Due=$863.08, Claims Funding=$1,502.29, Total (4)=$2,365.37 → current_premium="$2,365.37" ✅ NOT "$863.08" ❌

🔹 BLUECROSS BLUE SHIELD (BCBS) / ANTHEM SPECIAL RULES (CRITICAL):
- **ALWAYS extract the value from the "Total Premium" column as the `current_premium`.**
- **STRICTLY FORBIDDEN:** Do NOT use the "Employee Medical", "Dependent(s) Medical", or "Total Medical" columns for the premium value on these forms. 
- You MUST ignore the internal sub-columns and jump to the far-right "Total Premium" column for every member.
- **PHASE 1 EXCLUSION — "Eligibility Changes" section:** This section (identified by the heading "Eligibility Changes" and rows with a "Change Code" column such as ADDSUB or ADDDEP) represents mid-period enrollment events and is NOT part of the regular billing roster. **You MUST completely ignore every row inside the "Eligibility Changes" section. Do NOT extract any employee records from it.**

🔹 COVERAGE FALLBACK (e.g. BLUECROSS):
- If the document lacks an explicit 'Coverage' column or it is blank, YOU MUST INFER the coverage tier from the relationship or enrollee type (e.g. 'EE', 'Subscriber' -> EE, 'SP', 'Spouse' -> ES).

🔹# CORRECTED UHC PREMIUM SELECTION RULE

# CASE 1: Employee has PAIRED lines (Max Claims Liability + Admin/Excess Loss 
# under the SAME plan group/policy)
#   → current_premium = the shared "Total" column value (e.g., $742.11)
#   → Apply the SAME Total to BOTH rows for that employee
#   → Do NOT use individual Charge Amounts ($399.09 / $343.02)

# CASE 2: Employee has a SINGLE line (no paired Admin/Excess Loss line)
#   → current_premium = the "Total" column value (same as before)

# CASE 3: Adjustment rows (negative values like -$742.11 for TRM status)
#   → Map to adjustment_amount, NOT current_premium
#   → current_premium = NULL for terminated rows

# FORBIDDEN: Never use raw Charge Amount as current_premium
#            when a Total column value is present for that employee group
🔹 Coverage Recovery  : Map coverage type codes using the following legend for ALL UHC documents:
    - `E` or "Employee Only" → **EE**
    - `ES` or "Employee and Spouse" → **ES**
    - `ES` or "Employee + Spouse" → **ES**
    - `ESC` or "Employee and Family" → **FAM**
    - `ESC` or "Employee + Family" → **FAM**
    - `EC` or "Employee and Child(ren)" → **EC**
    - `EC` or "Employee + Child(ren)" → **EC**
    - `E1D` or "Employee and One Dependent" → **EC**
    - `E2D` or "Employee and Two Dependents" → **EC**
    - `E3D` or "Employee and Three Dependents" → **EC**
    - `E4D` or "Employee and Four Dependents" → **EC**
    - `E5D` or "Employee & One or More Dependent" → **EC**
    - `E6D` or "Employee & Two or More Dependents" → **EC**
    - `E7D` or "Employee & Three or More Dependents" → **EC**
    - `E8D` or "Employee & Four or More Dependents" → **EC**
    - `E9D` or "Employee & Five or More Dependents" → **EC**
    - Single-letter codes only (when alone): `E` → **EE**, `S` → **ES**, `F` → **FAM**, `C` → **EC**, `E E` → **EE**

{
  "summary": {"company_name": "", "total_amount_due": ""},
  "employees": [{"full_name": null, "first_name": null, "middal_name": null, "last_name": null, "coverage": null, "plan_name": null, "plan_type": null, "current_premium": null, "adjustment_amount": null, "birth_date": null, "gender": null, "home_zip_code": null, "billing_period": null}]
}

🔹 SPECIAL CASE - PAYROLL / DEDUCTION REGISTERS (e.g. WARWICK / BENEFIT DEDUCTION ROSTER):
- If column headers include "Pay Date", "Deduction Date", or "Check Date", you MUST extract this into the `billing_period` field for EVERY row.
- **BENEFIT DEDUCTION ROSTER SPECIFIC (CRITICAL):** If the document is titled "BENEFIT DEDUCTION ROSTER", you MUST extract the value from the "Monthly Premium" -> "Total" column as the `current_premium`. 
- **STRICTLY FORBIDDEN:** Do NOT use the "Pay Period Amount" column values for the premium. Always use the monthly total.

PDF TEXT: {text}
"""
    # Page-based chunking with SMART OVERLAP to prevent data loss at page boundaries
    pages = split_text_by_pages(text)
    chunks = []
    
    if pages:
        PAGES_PER_CHUNK = 4  # Increased from 3 to 4 pages per chunk (reduces total chunks)
        OVERLAP_PAGES = 2    # 2-page overlap to handle page-boundary splits
        
        # Sliding window with 2-page overlap to ensure orphaned lines are captured
        # Example with 4-page chunks: [1,2,3,4], [3,4,5,6], [5,6,7,8]...
        # This gives same protection but ~30% fewer chunks than [1,2,3], [2,3,4], [3,4,5]...
        for i in range(0, len(pages), PAGES_PER_CHUNK - OVERLAP_PAGES):
            chunk_pages = pages[i : i + PAGES_PER_CHUNK]
            if chunk_pages:
                chunks.append("\n".join(chunk_pages))
            if i + PAGES_PER_CHUNK >= len(pages):
                break

    all_employees = []
    final_summary = {}

    def _call_llm_for_chunk(chunk_text: str) -> dict | None:
        """Send one chunk to GPT-4o and return parsed dict, or None on failure."""
        prompt = prompt_template.replace("{text}", chunk_text)
        try:
            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": "You are a precise insurance billing data extraction assistant. Return valid JSON only. No markdown. No extra text."},
                    {"role": "user", "content": prompt},
                ],
                temperature=0,
                response_format={"type": "json_object"},
            )
            raw = response.choices[0].message.content

            # Strategy 1: Direct parsing
            try:
                return json.loads(raw)
            except json.JSONDecodeError:
                pass

            # Strategy 2: Remove markdown code blocks
            try:
                cleaned = re.sub(r"```json|```", "", raw).strip()
                return json.loads(cleaned)
            except json.JSONDecodeError:
                pass

            # Strategy 3: Extract JSON object from surrounding text
            try:
                json_match = re.search(r'\{[\s\S]*\}', raw)
                if json_match:
                    return json.loads(json_match.group())
            except json.JSONDecodeError:
                pass

            return None  # All strategies failed
        except Exception as e:
            print(f"    [RPVE] LLM call error: {e}")
            return None

    def _process_chunk(chunk_text: str, label: str, depth: int = 0) -> tuple[list, dict]:
        """
        Process a single chunk. If parsing fails, automatically split in half
        and retry each half independently (recursive).
        Returns (employees_list, summary_dict).
        """
        # If chunk is too small, don't try LLM, just return empty
        if not chunk_text.strip():
            return [], {}
        
        # Apply grouping PER CHUNK to handle page-boundary splits within chunks
        chunk_text_processed = chunk_text
        if sub_type in ["engage", "prestige", "velocity"]:
            chunk_text_processed = group_indented_lines(chunk_text)

        data = _call_llm_for_chunk(chunk_text_processed)
        if data is not None:
            return data.get("employees", []), data.get("summary", {})

        # If parsing failed and we haven't reached max depth/min size, split and retry
        if len(chunk_text) > 4000 and depth < 3:
            mid = len(chunk_text) // 2
            # Find the nearest newline to the midpoint to avoid splitting mid-record
            split_at = chunk_text.rfind('\n', 0, mid)
            if split_at == -1 or split_at < mid * 0.5: # Fallback if no good newline
                split_at = mid
            
            half_a = chunk_text[:split_at]
            half_b = chunk_text[split_at:]

            print(f"    [RPVE] {label} failed — retrying as sub-chunks (depth {depth+1})...")
            emps_a, summ_a = _process_chunk(half_a, f"{label}a", depth + 1)
            emps_b, summ_b = _process_chunk(half_b, f"{label}b", depth + 1)

            return emps_a + emps_b, summ_a or summ_b
        
        print(f"    [RPVE] {label} failed permanently (length {len(chunk_text)}).")
        return [], {}


    print(f"[RPVE] LLM extraction ({sub_type}) split into {len(chunks)} chunks...")
    
    # Process chunks in parallel to reduce total time
    with ThreadPoolExecutor(max_workers=min(len(chunks), 10)) as executor:
        futures = [
            executor.submit(_process_chunk, chunk, f"Chunk {i+1}/{len(chunks)}")
            for i, chunk in enumerate(chunks)
        ]
        
        for i, future in enumerate(futures):
            label = f"Chunk {i+1}/{len(chunks)}"
            try:
                emps, summ = future.result()
                if emps is not None:
                    if not final_summary and summ:
                        final_summary = summ
                    all_employees.extend(emps)
                    print(f"  [RPVE] {label} processed -> found {len(emps)} records")
                else:
                    print(f"  [RPVE] {label} failed: Could not parse JSON from response")
            except Exception as e:
                print(f"  [RPVE] {label} error: {e}")


    result = {
        "summary": final_summary,
        "employees": all_employees
    }
    
    # Special fallback for engage type when LLM extraction fails or returns empty
    if sub_type == "engage" and (not all_employees or len(all_employees) == 0):
        print(f"[RPVE] LLM extraction returned empty for engage, trying fallback extractors...")
        
        # Check if this is an ACSA-style summary invoice
        if "ASSOCIATION OF COMMUNITY SERVICE" in text.upper() or "ACSA" in text.upper():
            print(f"[RPVE] Detected ACSA format, using summary fallback...")
            fallback_result = extract_acsa_summary_fallback(text)
            if fallback_result and fallback_result.get("employees"):
                print(f"[RPVE] ACSA fallback successful, returning {len(fallback_result['employees'])} records")
                return fallback_result
        
        # Try generic regex extraction for other engage formats
        print(f"[RPVE] Trying generic engage fallback...")
        generic_fallback = extract_engage_fallback(text)
        if generic_fallback and generic_fallback.get("employees"):
            print(f"[RPVE] Generic engage fallback successful, returning {len(generic_fallback['employees'])} records")
            return generic_fallback
    
    return result


def deduplicate_employees(employees: list[dict]) -> list[dict]:
    """
    Removes duplicate employee records from a list using a two-pass strategy.

    Pass 1 — Exact deduplication:
        Removes records that are byte-for-byte identical on (full_name, plan_name,
        premium, adjustment, billing_period). Handles same-record repeated exactly.

    Pass 2 — Anomaly deduplication (new rule):
        If two records share the same first_name + last_name + current_premium,
        they are considered the same person extracted twice (e.g. from overlapping
        page chunks with slightly different plan names or coverage codes).
        The record with the most populated fields is kept; the other is dropped.

    Args:
        employees: A list of employee data dictionaries extracted from an invoice.

    Returns:
        A deduplicated list of employee data dictionaries.
    """
    original_count = len(employees)

    # ── PASS 1: Exact key deduplication ───────────────────────────────────────
    seen = set()
    pass1_list = []
    for employee in employees:
        plan_name = (employee.get("plan_name") or "").strip()
        full_name = (employee.get("full_name") or "").strip()

        if not full_name:
            first_name = (employee.get("first_name") or "").strip()
            last_name  = (employee.get("last_name") or "").strip()
            if first_name and last_name:
                full_name = f"{first_name} {last_name}"

        # If name or plan is missing, keep the record unconditionally
        if not full_name or not plan_name:
            pass1_list.append(employee)
            continue

        premium        = str(employee.get("current_premium") or "").strip()
        adjustment     = str(employee.get("adjustment_amount") or "").strip()
        billing_period = str(employee.get("billing_period") or "").strip()
        unique_key = (full_name.upper(), plan_name.upper(), premium, adjustment, billing_period)
        if unique_key not in seen:
            seen.add(unique_key)
            pass1_list.append(employee)

    # ── PASS 2: Name + Premium anomaly deduplication ──────────────────────────
    # If two records share the same first_name, last_name AND current_premium,
    # AND their plan names are compatible (same, or one is a substring of the other),
    # they are considered the same person extracted twice from overlapping page chunks.
    # Keep the record with the most populated (non-null/non-empty) fields.
    #
    # ⚠ Safety guard: if plan names are completely different (e.g. "Medical" vs "Dental")
    # we do NOT collapse them — they are legitimately distinct plan rows.

    def _field_score(emp: dict) -> int:
        """Score a record by how many fields are non-null/non-empty."""
        score = sum(1 for v in emp.values() if v is not None and str(v).strip())
        # Bonus: prefer longer (more complete) plan names
        score += len(str(emp.get("plan_name") or "")) // 10
        # Bonus: prefer standard normalized coverage codes
        std_coverages = {"EE", "ES", "FAM", "EC"}
        if str(emp.get("coverage") or "").strip().upper() in std_coverages:
            score += 1
        return score

    def _plans_compatible(plan_a: str, plan_b: str) -> bool:
        """Return True if two plan names are the same or one is a substring of the other."""
        a = plan_a.strip().upper()
        b = plan_b.strip().upper()
        if not a or not b:
            return True   # one is blank — treat as compatible (no conflict)
        if a == b:
            return True
        # Substring match handles wrapped/truncated plan names across page chunks
        if len(a) > 3 and len(b) > 3:
            if a in b or b in a:
                return True
        return False

    grouped: dict = {}
    order: list = []
    for emp in pass1_list:
        first_name = str(emp.get("first_name") or "").strip().upper()
        last_name  = str(emp.get("last_name") or "").strip().upper()
        premium    = str(emp.get("current_premium") or "").strip()
        plan_name  = str(emp.get("plan_name") or "").strip()

        # Only group if all three key fields are present
        if not first_name or not last_name or not premium:
            uid = id(emp)
            grouped[uid] = [emp]
            order.append(uid)
            continue

        # Find an existing group whose plan name is compatible (not a different plan)
        matched_key = None
        for existing_key in list(grouped.keys()):
            if not isinstance(existing_key, tuple):
                continue
            if existing_key[:2] != (first_name, last_name) or existing_key[2] != premium:
                continue
            # Check plan name compatibility against first record in that group
            existing_plan = str(grouped[existing_key][0].get("plan_name") or "")
            if _plans_compatible(plan_name, existing_plan):
                matched_key = existing_key
                break

        if matched_key is not None:
            grouped[matched_key].append(emp)
        else:
            # New unique key (different plan or first occurrence)
            name_premium_key = (first_name, last_name, premium)
            # Make key unique if same name+premium but different plan (e.g. Medical vs Dental)
            disambiguated_key = name_premium_key
            suffix = 0
            while disambiguated_key in grouped:
                suffix += 1
                disambiguated_key = (first_name, last_name, premium, suffix)
            grouped[disambiguated_key] = [emp]
            order.append(disambiguated_key)

    final_list = []
    for key in order:
        candidates = grouped[key]
        if len(candidates) == 1:
            final_list.append(candidates[0])
        else:
            # Multiple records with same name + premium + compatible plan → keep the best
            best = max(candidates, key=_field_score)
            final_list.append(best)
            removed = len(candidates) - 1
            name_label = f"{key[0]} {key[1]}" if isinstance(key, tuple) else str(key)
            prem_label = key[2] if isinstance(key, tuple) and len(key) >= 3 else '?'
            print(f"[RPVE] Anomaly dedup: kept best record for '{name_label}' (premium={prem_label}), dropped {removed} weaker duplicate(s)")

    final_count   = len(final_list)
    total_removed = original_count - final_count
    if total_removed > 0:
        print(f"[RPVE] Deduplication: {original_count} rows -> {final_count} rows ({total_removed} total duplicates removed)")

    return final_list

def build_excel(data: dict, sub_type: str, stem: str, active_employee_fields: list[str] | None = None, out_dir: Path | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    out = out_dir if out_dir is not None else OUTPUT_DIR
    wb        = Workbook()
    hex_col   = HEADER_COLOURS.get(sub_type, "1A1A2E")
    hdr_fill  = PatternFill("solid", fgColor=hex_col)
    hdr_font  = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="DDDDDD")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)
    da        = Alignment(vertical="center")
    tf        = Font(bold=True, size=11, name="Calibri")
    tfill     = PatternFill("solid", fgColor="F0F0F0")

    summary   = data.get("summary", {})
    employees = data.get("employees", [])

    # ── Sheet 1: Employee Details ─────────────────────────────────────────────
    we = wb.active
    we.title = "Employee Details"
    we.sheet_view.showGridLines = False

    # Use passed fields or fallback to global mapping
    all_cols = active_employee_fields if active_employee_fields is not None else EMPLOYEE_FIELDS.get(sub_type, [])

    we.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(all_cols),1))
    t2 = we.cell(row=1, column=1, value=f"Employee Details - {len(employees)} records")
    t2.font      = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    t2.fill      = PatternFill("solid", fgColor=hex_col)
    t2.alignment = Alignment(horizontal="center", vertical="center")
    we.row_dimensions[1].height = 28

    for ci, col in enumerate(all_cols, 1):
        c = we.cell(row=2, column=ci, value=col.replace("_", " "))
        c.fill, c.font, c.alignment, c.border = hdr_fill, hdr_font, hdr_align, bdr
        we.column_dimensions[get_column_letter(ci)].width = 22
    we.row_dimensions[2].height = 22

    for ri, emp in enumerate(employees, 3):
        we.row_dimensions[ri].height = 18
        for ci, col in enumerate(all_cols, 1):
            c = we.cell(row=ri, column=ci, value=emp.get(col.lower(), ""))
            c.border, c.alignment = bdr, da
            c.font = Font(size=10, name="Calibri")
            if ri % 2 == 0:
                c.fill = PatternFill("solid", fgColor="F7F7F7")

    fin_cols = {"CURRENT_PREMIUM", "MONTHLY_TOTAL_PREMIUM", "GRAND_TOTAL", "TOTAL_COST", "ADJUSTMENT_AMOUNT"}
    fin_present = [c for c in all_cols if c.upper() in fin_cols]
    if fin_present and employees:
        tr = len(employees) + 3
        we.row_dimensions[tr].height = 20
        lc = we.cell(row=tr, column=1, value="TOTAL")
        lc.font, lc.fill, lc.border = tf, tfill, bdr
        for ci, col in enumerate(all_cols, 1):
            c = we.cell(row=tr, column=ci)
            c.fill, c.border = tfill, bdr
            if col.upper() in fin_cols:
                total = 0.0
                for emp in employees:
                    # Skip subtotal rows ("TOTAL" or empty plan_name) to avoid double counting
                    p_opt = emp.get("coverage_option", "")
                    pname = str(p_opt if p_opt is not None else "").strip().upper()
                    
                    if pname == "TOTAL" or (sub_type in ["engage", "velocity"] and (not pname or pname == "NONE")):
                        continue
                    
                    val_str = str(emp.get(col.lower(), "")).replace("$", "").replace(",", "")
                    try:
                        clean_val = re.sub(r'[^\d.-]', '', val_str)
                        if clean_val:
                            total += float(clean_val)
                    except:
                        pass
                c.value = f"${total:,.2f}"
                c.font  = tf

    we.freeze_panes = "A3"
    wb.active = we

    xlsx_path = out / f"{stem}_RPVE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(str(xlsx_path))
    _FILE_REGISTRY[xlsx_path.name] = str(xlsx_path.resolve())  # register for instant download
    print(f"[RPVE] Excel -> {xlsx_path.name}")
    return xlsx_path


def build_json_file(data: dict, sub_type: str, stem: str, active_employee_fields: list[str] | None = None, out_dir: Path | None = None) -> Path:
    out = out_dir if out_dir is not None else OUTPUT_DIR
    summary   = data.get("summary", {})
    employees = data.get("employees", [])
    required  = active_employee_fields if active_employee_fields is not None else EMPLOYEE_FIELDS.get(sub_type, [])

    rows = []
    for emp in employees:
        row = {}
        for k, v in summary.items():
            row[k.upper()] = emp.get(k.lower(), v)
        # Only include required fields - strip everything else
        for col in required:
            row[col] = emp.get(col.lower(), "")
        row["RPVE_SUB_TYPE"] = sub_type.upper()
        rows.append(row)

    json_path = out / f"{stem}_RPVE_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(str(json_path), "w", encoding="utf-8") as f:
        json.dump(rows, f, indent=2, ensure_ascii=False)
    _FILE_REGISTRY[json_path.name] = str(json_path.resolve())  # register for instant download
    print(f"[RPVE] JSON  -> {json_path.name}")
    return json_path

# ══════════════════════════════════════════════════════════════════════════════
# FASTAPI APP
# ══════════════════════════════════════════════════════════════════════════════


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Start background worker pool and recover stale jobs on startup."""
    import job_store
    import job_worker
    recovered = job_store.recover_stale_jobs()
    if recovered:
        print(f"[RPVE] Recovered {recovered} stale job(s) from previous run.")
    job_worker.start_workers()
    print("[RPVE] Background worker pool started.")
    yield
    # Shutdown: nothing to tear down — worker threads are daemons.


app = FastAPI(
    title="RPVE - Benefit Invoice Extractor",
    description="Resourcing · Prestige · Velocity · Engage",
    version="1.0.0",
    lifespan=lifespan,
)

from fastapi.openapi.utils import get_openapi

def custom_openapi():
    if app.openapi_schema:
        return app.openapi_schema
    
    openapi_schema = get_openapi(
        title="RPVE - Benefit Invoice Extractor",
        version="1.0.0",
        routes=app.routes,
    )
    
    # 1. Fix components/schemas
    for component in openapi_schema.get("components", {}).get("schemas", {}).values():
        properties = component.get("properties", {})
        if not isinstance(properties, dict):
            continue
        for prop in properties.values():
            if not isinstance(prop, dict):
                continue
            if prop.get("contentMediaType") == "application/octet-stream":
                prop["format"] = "binary"
                del prop["contentMediaType"]
            # Array of files
            if prop.get("type") == "array" and isinstance(prop.get("items"), dict):
                items = prop["items"]
                if items.get("contentMediaType") == "application/octet-stream":
                    items["format"] = "binary"
                    del items["contentMediaType"]

    # 2. Fix request bodies in paths (inline schemas in newer FastAPI)
    for path_data in openapi_schema.get("paths", {}).values():
        for operation in path_data.values():
            if not isinstance(operation, dict):
                continue
            request_body = operation.get("requestBody")
            if not isinstance(request_body, dict):
                continue
            content = request_body.get("content", {})
            form_data = content.get("multipart/form-data", {})
            schema = form_data.get("schema", {})
            properties = schema.get("properties", {})
            if not isinstance(properties, dict):
                continue
            for prop in properties.values():
                if not isinstance(prop, dict):
                    continue
                if prop.get("type") == "string" and prop.get("contentMediaType") == "application/octet-stream":
                    prop["format"] = "binary"
                    del prop["contentMediaType"]
                elif prop.get("type") == "array" and isinstance(prop.get("items"), dict):
                    items = prop["items"]
                    if items.get("type") == "string" and items.get("contentMediaType") == "application/octet-stream":
                        items["format"] = "binary"
                        del items["contentMediaType"]

    app.openapi_schema = openapi_schema
    return openapi_schema

app.openapi = custom_openapi

app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

# ── Global file registry: filename → absolute path ───────────────────────────
# Populated whenever build_excel / build_json_file writes a file.
# Download endpoint checks this first — O(1), no disk scan at all.
_FILE_REGISTRY: dict[str, str] = {}

FRONTEND_DIST_DIR = BASE_DIR / "frontend" / "dist"
(FRONTEND_DIST_DIR / "assets").mkdir(parents=True, exist_ok=True)
app.mount("/assets", StaticFiles(directory=FRONTEND_DIST_DIR / "assets"), name="assets")

@app.get("/", response_class=HTMLResponse, include_in_schema=False)
async def serve_ui():
    ui = FRONTEND_DIST_DIR / "index.html"
    if ui.exists():
        return HTMLResponse(content=ui.read_text(encoding="utf-8"))
    
    ui_fallback = BASE_DIR / "RPVE_ui.html"
    if ui_fallback.exists():
        return HTMLResponse(content=ui_fallback.read_text(encoding="utf-8"))
        
    return HTMLResponse("<h2>RPVE running</h2><p>Build the frontend first.</p><a href='/docs'>Swagger -></a>")


@app.get("/api/health")
async def health():
    return {"status": "ok", "service": "RPVE", "sub_types": list(KEYWORDS.keys())}


def forward_fill_plan_continuations(text: str) -> str:
    """
    Scans the invoice text page-by-page and identifies plan name headers.
    If a subsequent page/section has '(cont.)' instead of the repeated plan name,
    replaces '(cont.)' with the active plan name so that LLM chunking (which receives 
    only a subset of pages) has the necessary plan name context.
    """
    if not text:
        return text

    # Helper to check if a line is part of a table header
    def is_header_line(line: str) -> bool:
        l = line.lower()
        return any(kw in l for kw in ["empl", "eff", "medical", "total", "name", "date", "*type", "amount", "current", "inforce", "charges", "retroactivity"])

    # Helper to check if a line is a premium sub-row rather than a plan name
    def is_sub_row(line: str) -> bool:
        l_strip = line.strip()
        # Matches e.g., "0520 277.52" or "0106 0.00 $480.23"
        if re.match(r'^\d{4}\s+[\d.,\s$]+', l_strip):
            return True
        return False

    # Split text by page markers
    pages = re.split(r'(--- Page \d+ ---)', text)
    reconstructed = []
    active_plan_lines = []

    for part in pages:
        if not part.strip() or part.startswith("--- Page"):
            reconstructed.append(part)
            continue
            
        lines = part.split("\n")
        new_lines = []
        
        header_seen = False
        collecting_plan = False
        plan_lines = []
        
        for line in lines:
            l_strip = line.strip()
            if not l_strip:
                new_lines.append(line)
                continue
                
            # Detect section/table headers
            is_header = (
                "empl" in l_strip.lower() and "name" in l_strip.lower()
            ) or (
                "current" in l_strip.lower() and "charges" in l_strip.lower()
            ) or (
                "retroactivity" in l_strip.lower() and "charges" in l_strip.lower()
            )
            
            if is_header:
                header_seen = True
                collecting_plan = True
                plan_lines = []
                new_lines.append(line)
                continue
                
            if collecting_plan:
                # Employee lines contain commas (names) and digits (dates/amounts)
                is_employee = "," in l_strip and any(char.isdigit() for char in l_strip)
                
                if is_employee:
                    collecting_plan = False
                    if plan_lines:
                        cleaned_plan = [pl for pl in plan_lines if not is_header_line(pl) and not is_sub_row(pl) and "(cont" not in pl.lower() and pl.strip()]
                        if cleaned_plan:
                            active_plan_lines = cleaned_plan
                    new_lines.append(line)
                else:
                    if "(cont" in l_strip.lower():
                        if active_plan_lines:
                            # Replace (cont.) line with the active plan lines
                            for apl in active_plan_lines:
                                new_lines.append(apl)
                        else:
                            new_lines.append(line)
                    else:
                        plan_lines.append(line)
                        new_lines.append(line)
            else:
                if "(cont" in l_strip.lower() and active_plan_lines:
                    for apl in active_plan_lines:
                        new_lines.append(apl)
                else:
                    new_lines.append(line)
                    
        reconstructed.append("\n".join(new_lines))

    return "".join(reconstructed)


async def process_invoice_data(file_path: Path, original_filename: str, out_dir: Path | None = None):
    print(f"\n[RPVE] Processing -> {original_filename}")
    ext = Path(original_filename).suffix.lower()
    
    # ── UNIQUE OUTPUT PATH ────────────────────────────────────────────
    # Clean and truncate stem for the output folder (cosmetic naming)
    safe = re.sub(r'[\\/:*?"<>|]', "_", original_filename)
    stem = Path(safe).stem.replace(" ", "_").strip()[:50]
    
    # Use provided out_dir (job-scoped) or fall back to timestamped OUTPUT_DIR sub-dir
    if out_dir is not None:
        run_out_dir = out_dir
    else:
        timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        run_out_dir = OUTPUT_DIR / f"{stem}_{timestamp_str}"
    run_out_dir.mkdir(parents=True, exist_ok=True)


    try:
        # Using local extract_text function instead of missing identification module
        text = extract_text(file_path)
        
        # Forward-fill plan names when they are continued across pages/sections
        text = forward_fill_plan_continuations(text)
        
        # Consistent Text Output: Save the extracted text for ALL file types
        txt_path = run_out_dir / f"{stem}.txt"
        with open(txt_path, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"[RPVE] Saved structured text to {txt_path.name}")
        
    except Exception as read_err:
        print(f"[RPVE] Read error: {read_err}")
        raise Exception(f"Failed to extract text from {ext} file: {read_err}")

    if not text.strip():
        raise Exception("No text extracted. File may be empty or an unreadable image.")

    # ---- Filename-based fallback for Data Link EMI ----
    # Normalise filename: replace underscores/hyphens with spaces so that
    # "Data_Link_EMI_Invoice..." and "Data-Link-EMI..." both match correctly.
    safe_normalised = safe.upper().replace("_", " ").replace("-", " ")
    if "DATA LINK EMI" in safe_normalised:
        sub_type = "datalink_emi"
        print(f"[RPVE] Filename -> Data Link EMI detected, forcing sub_type. (key: {safe_normalised[:40]})")
    else:
        # Use classification to determine carrier sub-type
        sub_type = classify(text)
    print(f"[RPVE] Classified as -> {sub_type.upper()}")

    try:
        data = extract_with_llm(sub_type, text)
    except Exception as e:
        raise Exception(f"LLM extraction failed: {str(e)}")

    # Clean up results: Remove rows without names (Center for Human Development fix)
    all_emps = data.get("employees", [])
    valid_emps = []
    for e in all_emps:
        # Require at least one name field to be present
        fname = str(e.get("first_name") or "").strip()
        lname = str(e.get("last_name") or "").strip()
        fulln = str(e.get("full_name") or "").strip()
        if (fname and lname) or fulln:
            valid_emps.append(e)
    data["employees"] = valid_emps
    emp_count = len(valid_emps)
    print(f"[RPVE] Extracted  -> {emp_count} rows")

    # ── Global FULL_NAME construction (applies to ALL file types) ─────────────
    # If the LLM did not return a full_name, build it from first/middle/last.
    for emp in data["employees"]:
        if not str(emp.get("full_name") or "").strip():
            parts = [
                str(emp.get("first_name") or "").strip(),
                str(emp.get("middal_name") or "").strip(),
                str(emp.get("last_name") or "").strip(),
            ]
            emp["full_name"] = " ".join(p for p in parts if p)
            
        # ── Plan Name: NO fallback to plan_type ───────────────────────────────
        # If the invoice has no explicit plan name column, plan_name stays null.
        # Do NOT copy plan_type into plan_name — that would be hallucination.

    # ── Post-LLM clean-up for Data Link EMI ─────────────────────────────────
    if sub_type == "datalink_emi":
        medical_lookup: dict[str, str] = {}
        retro_lookup:   dict[str, str] = {}
        try:
            import pdfplumber
            with pdfplumber.open(str(file_path)) as _pdf:
                for _page in _pdf.pages:
                    _tables = _page.extract_tables()
                    for _tbl in _tables:
                        for _row in _tbl:
                            if not _row or len(_row) < 5: continue
                            _name_cell = str(_row[1] or "").strip()
                            if not _name_cell or _name_cell.lower() in ("name", ""): continue
                            _is_retro = False
                            _medical_idx = 3
                            if len(_row) >= 9:
                                _date_cell = str(_row[2] or "")
                                if "/" in _date_cell or "-" in _date_cell:
                                    _is_retro = True
                                    _medical_idx = 4
                            _medical_val = str(_row[_medical_idx] or "").strip()
                            if not re.match(r'^\$?-?[\d,]+\.\d{2}$', _medical_val.replace("(", "").replace(")", "")): continue
                            _key = re.sub(r'\s+', ' ', _name_cell.upper().replace(",", "")).strip()
                            if _is_retro: retro_lookup[_key] = _medical_val
                            else: medical_lookup[_key] = _medical_val
            print(f"[RPVE] Data Link EMI: built lookup tables.")
        except Exception as _lookup_err:
            print(f"[RPVE] Data Link EMI: pdfplumber lookup failed ({_lookup_err})")

        def _emi_key(emp: dict) -> str:
            fn = str(emp.get("first_name") or "").strip().upper()
            ln = str(emp.get("last_name") or "").strip().upper()
            return f"{ln} {fn}" if fn and ln else ""

        for emp in data.get("employees", []):
            _k = _emi_key(emp)
            _is_retro_emp = emp.get("adjustment_amount") and not emp.get("current_premium")
            if _is_retro_emp:
                if _k and _k in retro_lookup: emp["adjustment_amount"] = retro_lookup[_k]
                emp["current_premium"] = None
            else:
                if _k and _k in medical_lookup: emp["current_premium"] = medical_lookup[_k]
                emp["adjustment_amount"] = None

    data["employees"] = deduplicate_employees(data["employees"])
    analysis_file_name = None

    try:
        active_fields = EMPLOYEE_FIELDS.get(sub_type, UNIFIED_FIELDS)
        extracted_text_upper = text.upper()
        is_strict_adp = ("TOTALSOURCE" in extracted_text_upper or "ADP" in extracted_text_upper or "NCT3-EPO" in extracted_text_upper)
        is_peo = is_strict_adp or "INSPERITY" in extracted_text_upper
        
        analysis_data = []
        if is_peo:
            from collections import defaultdict
            grouped2: dict = defaultdict(list)
            for emp in data.get("employees", []):
                pname = str(emp.get("plan_name") or "").strip().upper()
                ptype = str(emp.get("plan_type") or "").strip().upper()
                copt  = str(emp.get("coverage_option") or "").strip().upper()
                if any(x in pname or x in ptype or x in copt for x in ("TOTAL", "SUBTOTAL", "GRAND TOTAL")): continue
                key = (
                    str(emp.get("first_name", "")).strip().upper(),
                    str(emp.get("last_name", "")).strip().upper(),
                    str(emp.get("company_name", "")).strip().upper()
                )
                grouped2[key].append(emp)

            collapsed = []
            for (fname, lname, company_name), rows in grouped2.items():
                if not rows: continue
                parsed_rows = []
                for r in rows:
                    val_str = str(r.get("current_premium") or "").replace("$", "").replace(",", "")
                    try: v = round(float(re.sub(r'[^\d.-]', '', val_str)), 2)
                    except: v = 0.0
                    parsed_rows.append((v, r))
                parsed_rows.sort(key=lambda x: x[0], reverse=True)
                valid_benefit_rows = []
                if len(parsed_rows) > 1:
                    top_val, top_row = parsed_rows[0]
                    remaining_sum = sum(v for v, _ in parsed_rows[1:])
                    if abs(top_val - remaining_sum) < 0.1: valid_benefit_rows = parsed_rows[1:]
                    else: valid_benefit_rows = parsed_rows
                else: valid_benefit_rows = parsed_rows
                if valid_benefit_rows:
                    valid_benefit_rows.sort(key=lambda x: x[0], reverse=True)
                    best_val, best_row = valid_benefit_rows[0]
                    pname = str(best_row.get("plan_name") or "").strip().upper()
                    if pname in ("TOTAL", "SUBTOTAL"): continue
                    cov_opt = str(best_row.get("coverage_option") or "").strip()
                    cat_name = str(best_row.get("plan_name") or "").strip()
                    if cov_opt and cov_opt.upper() not in ("TOTAL", "NONE", ""):
                        best_row = dict(best_row)
                        best_row["plan_name"] = cov_opt
                        if not best_row.get("plan_type"): best_row["plan_type"] = cat_name
                    if not str(best_row.get("full_name") or "").strip():
                        best_row = dict(best_row)
                        parts = [str(best_row.get("first_name") or "").strip(), str(best_row.get("middal_name") or "").strip(), str(best_row.get("last_name") or "").strip()]
                        best_row["full_name"] = " ".join(p for p in parts if p)
                    collapsed.append(best_row)
            data["employees"] = collapsed

        is_uhc = any(x in extracted_text_upper for x in ["UNITEDHEALTHCARE", "UNITED HEALTHCARE", "UHC"])
        is_excel = ext in [".xlsx", ".xls"]
        final_employees = []
        for emp in data.get("employees", []):
            val_str = str(emp.get("current_premium") or "").replace("$", "").replace(",", "")
            try: premium_val = round(float(re.sub(r'[^\d.-]', '', val_str)), 2)
            except: premium_val = 0.0
            # For PEO/ADP documents, skip the $250 filter — the PEO collapse already
            # selected the single best plan row per employee (typically their medical plan).
            # Applying $250 here would silently drop employees whose best extracted row
            # happened to be Dental/Vision (e.g. when LLM missed the medical line on a page).
            # For all other document types, keep the $250 analysis split.
            if is_peo:
                final_employees.append(emp)
            elif premium_val < 250:
                analysis_data.append(emp)
            else:
                final_employees.append(emp)
        data["employees"] = final_employees

        xlsx_path = build_excel(data, sub_type, stem, active_employee_fields=active_fields, out_dir=run_out_dir)
        json_path = build_json_file(data, sub_type, stem, active_employee_fields=active_fields, out_dir=run_out_dir)
        if analysis_data:
            analysis_path = run_out_dir / f"{stem}_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(str(analysis_path), "w", encoding="utf-8") as af:
                json.dump(analysis_data, af, indent=2, ensure_ascii=False)
            analysis_file_name = analysis_path.name
            # NOTE: No _cache write — files are found via rglob on download.

    except Exception as build_err:
        import traceback
        print(f"[RPVE] Output building error:\n{traceback.format_exc()}")
        raise Exception(f"Failed to generate output files: {str(build_err)}")

    # NOTE: _cache writes removed — output files live in run_out_dir which is
    # either OUTPUT_DIR (legacy /api/extract) or a job-scoped work dir.

    summary_dict = data.get("summary", {})
    total_val_str = "0"
    for tk in ["total_cost", "grand_total", "total_amount_due", "total_balance_due", "amount_due", "total_amount"]:
        val = summary_dict.get(tk) or summary_dict.get(tk.upper())
        if val:
            total_val_str = val
            break
    try: numeric_total = float(re.sub(r'[^0-9\.]', '', str(total_val_str)))
    except: numeric_total = 0.0

    return {
        "status": "success", "type": "INVOICE", "sub_type": sub_type, "sub_type_label": "Standard Mode",
        "employee_count": emp_count, "fields_in_excel": active_fields, "summary": summary_dict,
        "excel_file": xlsx_path.name, "json_file": json_path.name, 
        "excel_path": str(xlsx_path.absolute()), "json_path": str(json_path.absolute()),
        "output_file": xlsx_path.name,
        "output_json": json_path.name, "total_value": numeric_total,
        "excel_url": f"/api/download/{xlsx_path.name}", "json_url": f"/api/download/{json_path.name}",
        "analysis_file": analysis_file_name, "analysis_url": f"/api/download/{analysis_file_name}" if analysis_file_name else None,
        "employees": [{col: emp.get(col.lower(), "") for col in active_fields} for emp in data.get("employees", [])],
    }

# ── Sync wrapper used by flow_orchestrator (runs in worker thread) ────────────
def process_invoice_data_sync(file_path: Path, original_filename: str, out_dir: Path | None = None) -> dict:
    """
    Synchronous version of process_invoice_data for use in background threads.
    The async version is kept for the /api/extract endpoint (event-loop context).
    """
    import asyncio
    return asyncio.run(process_invoice_data(file_path, original_filename, out_dir=out_dir))


@app.post("/extract")
@app.post("/api/extract")
async def extract(file: UploadFile = File(...)):
    print(f"\n[RPVE] Extraction Mode -> Standard")
    if not file.filename:
        raise HTTPException(400, "No filename provided")
    ext = Path(file.filename).suffix.lower()
    if ext not in [".pdf", ".csv", ".xlsx", ".xls"]:
        raise HTTPException(400, f"Supported formats: PDF, CSV, XLSX, XLS. Got: {ext}")

    unique_id = uuid.uuid4().hex[:8]
    timestamp_str = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_filename = re.sub(r'[\\/:*?\"<>|]', '_', file.filename)
    filename_unique = f"{unique_id}_{timestamp_str}_{safe_filename}"
    file_path = UPLOAD_DIR / filename_unique
    
    with open(file_path, "wb") as f:
        shutil.copyfileobj(file.file, f)
    
    try:
        return await process_invoice_data(file_path, file.filename)
    except Exception as e:
        raise HTTPException(500, str(e))

@app.post("/process-flow")
@app.post("/api/process-flow")
async def process_flow(files: List[UploadFile] = File(...)):
    """
    Async, non-blocking implementation of the full RPVE pipeline.

    Frontend contract is UNCHANGED:
      - Still accepts the same multipart/form-data payload
      - Still blocks until the job completes (long-poll style)
      - Still returns the identical JSON response shape

    Internally, the job is executed in a background worker thread
    so the event loop is never blocked during OCR/LLM processing.
    Multiple simultaneous calls will each run in their own thread.
    The job_id is purely backend-internal — never sent to the frontend.
    """
    import job_store
    import job_worker

    print(f"\n[RPVE] Processing Flow with {len(files)} files")

    # ── 1. Generate a backend-internal job_id ─────────────────────────────────
    job_id  = uuid.uuid4().hex
    job_dir = job_worker.get_job_dir(job_id)
    input_dir = job_dir / "input"

    # ── 2. Save uploaded files into jobs/{job_id}/input/ ──────────────────────
    pdf_file    = None
    excel_files = []

    for file in files:
        if not file.filename:
            continue
        ext = Path(file.filename).suffix.lower()
        safe_filename   = re.sub(r'[\\/:*?\"<>|]', '_', file.filename)
        dest_path = input_dir / safe_filename
        # If same filename sent twice, suffix with a counter to avoid collision
        if dest_path.exists():
            dest_path = input_dir / f"{uuid.uuid4().hex[:4]}_{safe_filename}"

        with open(dest_path, "wb") as f:
            shutil.copyfileobj(file.file, f)

        if ext == ".pdf":
            pdf_file = dest_path
        elif ext in [".xlsx", ".xls"]:
            excel_files.append(dest_path)

    # Must have at least one Excel file (the template)
    if not excel_files:
        raise HTTPException(400, "Please upload at least one Excel template.")

    # Must have a source (either a PDF invoice or a second Excel file)
    if not pdf_file and len(excel_files) < 2:
        raise HTTPException(400, "Please upload either a PDF and an Excel, or at least two Excel files (Source + Template).")

    # ── 3. Create job record in DB and enqueue ────────────────────────────────
    job_store.create_job(job_id)
    job_worker.enqueue_job(job_id)
    print(f"[RPVE] Job {job_id[:8]}... enqueued.")

    # ── 4. Long-poll: wait for the job to complete (non-blocking) ─────────────
    # The event loop is free to serve other requests while we sleep.
    MAX_WAIT_SECONDS = 1200  # 20-minute hard limit per job (increased for larger PDFs with 2-page overlap)
    POLL_INTERVAL    = 1.0   # check every 1 second
    waited = 0.0

    while waited < MAX_WAIT_SECONDS:
        await asyncio.sleep(POLL_INTERVAL)
        waited += POLL_INTERVAL

        meta = job_store.get_job(job_id)
        if meta is None:
            raise HTTPException(500, "Job record lost — internal error")

        if meta.status == "completed":
            # Deserialise the rich result JSON that run_job() stored
            if meta.result_json:
                res_data = json.loads(meta.result_json)
                try:
                    from database import poc_db
                    poc_db.log_rpve_run(job_id, ", ".join([f.filename for f in files]), "SUCCESS", str(res_data.get("insurer", "")), str(res_data.get("total_value", "")))
                    print(f"[DB] Logged RPVE flow run for job {job_id[:8]} to converter.db", flush=True)
                except Exception as db_err:
                    print(f"[WARN] Failed to log RPVE flow run to DB: {db_err}", flush=True)
                return res_data
            raise HTTPException(500, "Job completed but result_json is empty")

        if meta.status == "failed":
            raise HTTPException(500, f"Processing failed: {meta.error or 'unknown error'}")

    # Timeout — mark as failed and return error
    job_store.update_status(job_id, "failed", error="Timed out after 10 minutes")
    raise HTTPException(504, "Processing timed out. Please try again.")


@app.get("/api/download/{filename}", include_in_schema=False)
async def download(filename: str, abs_path: str | None = None):
    """
    Serve a generated output file by name.

    Fast path: if the caller passes ?abs_path=<full_path>, the file is served
    directly without any rglob scan (O(1) lookup).  The path is validated to
    ensure it lives inside JOBS_DIR or OUTPUT_DIR before serving.

    Fallback (slow path — only used when abs_path is missing or invalid):
      1. Per-job work/output directories (jobs/{job_id}/**)
      2. Legacy OUTPUT_DIR subdirectories (rpve_outputs/**)
      3. Root of OUTPUT_DIR
    """
    ALLOWED_ROOTS = [JOBS_DIR.resolve(), OUTPUT_DIR.resolve()]

    def _serve(fp: Path) -> FileResponse:
        mt = (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            if filename.endswith(".xlsx")
            else "application/json"
        )
        return FileResponse(path=fp, filename=filename, media_type=mt)

    # ── Path 1 (fastest): in-memory registry lookup — O(1), no disk scan ─────
    if filename in _FILE_REGISTRY:
        cached = Path(_FILE_REGISTRY[filename])
        if cached.exists():
            return _serve(cached)
        else:
            # File was deleted externally — remove stale entry
            del _FILE_REGISTRY[filename]

    # ── Path 2: abs_path from frontend query param — single file check ────────
    if abs_path:
        try:
            candidate = Path(abs_path).resolve()
            # Security: only serve files within allowed directories
            if any(str(candidate).startswith(str(root)) for root in ALLOWED_ROOTS):
                if candidate.exists() and candidate.is_file():
                    _FILE_REGISTRY[filename] = str(candidate)  # cache for next time
                    return _serve(candidate)
        except Exception:
            pass  # fall through to rglob

    # ── Path 3 (slow fallback): rglob scan — only hits on server restart ──────
    # 1. Search all per-job directories
    if JOBS_DIR.exists():
        job_matches = list(JOBS_DIR.rglob(filename))
        if job_matches:
            _FILE_REGISTRY[filename] = str(job_matches[0].resolve())  # cache for next time
            return _serve(job_matches[0])

    # 2. Search legacy OUTPUT_DIR subdirectories
    output_matches = list(OUTPUT_DIR.rglob(filename))
    if output_matches:
        _FILE_REGISTRY[filename] = str(output_matches[0].resolve())  # cache for next time
        return _serve(output_matches[0])

    # 3. Root of OUTPUT_DIR
    fp = OUTPUT_DIR / filename
    if fp.exists():
        _FILE_REGISTRY[filename] = str(fp.resolve())  # cache for next time
        return _serve(fp)

    raise HTTPException(404, f"File not found: {filename}")


@app.get("/{filename}", include_in_schema=False)
async def serve_root_files(filename: str):
    if filename.startswith("api") or filename in ["docs", "openapi.json"]:
        raise HTTPException(404)
    file_path = FRONTEND_DIST_DIR / filename
    if file_path.exists() and file_path.is_file():
        return FileResponse(file_path)
    raise HTTPException(404)


if __name__ == "__main__":
    import uvicorn
    port = 8009
    if "--port" in sys.argv:
        try: port = int(sys.argv[sys.argv.index("--port") + 1])
        except: pass

    print("\n" + "="*50)
    print("  RPVE - Benefit Invoice Extractor")
    print("="*50)
    print(f"  UI      ->  http://localhost:{port}")
    print(f"  Swagger ->  http://localhost:{port}/docs")
    print(f"  Logs    ->  {BASE_DIR / 'service.log'}")
    print("="*50 + "\n")

    # Configure uvicorn to use our existing logging setup instead of its own
    log_config = {
        "version": 1,
        "disable_existing_loggers": False,
        "handlers": {},
        "loggers": {
            "uvicorn":        {"handlers": [], "level": "INFO", "propagate": True},
            "uvicorn.error":  {"handlers": [], "level": "INFO", "propagate": True},
            "uvicorn.access": {"handlers": [], "level": "INFO", "propagate": True},
        },
    }
    uvicorn.run(
        "RPVE_standalone:app",
        host="0.0.0.0",
        port=port,
        reload=True,
        reload_excludes=["rpve_uploads/*", "rpve_outputs/*", "*.log", "*.txt"],
        log_config=log_config,
    )