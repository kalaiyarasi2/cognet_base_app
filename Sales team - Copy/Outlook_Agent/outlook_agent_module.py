"""
outlook_agent_module.py
=========================================================================
Self-contained module for the Outlook Email Attachment Processing pipeline.
Import and drive from any external POC or script.

Public API
----------
    from outlook_agent_module import OutlookAgentModule

    # Instantiate (all params optional -- falls back to .env)
    agent = OutlookAgentModule()

    # One-shot run; returns a summary dict
    summary = agent.run(user_filter="Invoice")

    # Continuous polling
    agent.run_loop(user_filter="all", interval_seconds=30)

    # With a stored OAuth refresh_token (no device-code prompt)
    summary = agent.run(user_filter="all", refresh_token="<stored_token>")

    # Classify a single email (no full pipeline)
    category = agent.classify_email("Subject", "Body text")

    # Fetch raw unread emails only
    emails = agent.fetch_unread_emails()

    # Add a category at runtime (persisted to categories.json)
    agent.add_category("ACORD", "ACORD insurance forms and certificates")

Environment variables (.env fallbacks):
    AZURE_CLIENT_ID, AZURE_TENANT_ID, AZURE_CLIENT_SECRET
    OPENAI_API_KEY
    ONEDRIVE_FOLDER        (default: AI_Agent_Attachments)
    CATEGORIES_FILE        (default: ./config/categories.json)
    EXCEL_LOG_PATH         (default: ./logs/processed_mails_log.xlsx)
    PROCESSED_IDS_FILE     (default: ./config/processed_ids.json)

Dependencies:
    pip install msal openai langgraph requests python-dotenv openpyxl typing_extensions
=========================================================================
"""
from __future__ import annotations

import base64
import json
import logging
import os
import re
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Set

import msal
import requests as _requests
from dotenv import load_dotenv
from langgraph.graph import END, START, StateGraph
from openai import OpenAI
from typing_extensions import TypedDict

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Load .env
# ---------------------------------------------------------------------------
load_dotenv()

# ---------------------------------------------------------------------------
# Module logger
# ---------------------------------------------------------------------------
logger = logging.getLogger(__name__)
if not logger.handlers:
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(name)s -- %(message)s"))
    logger.addHandler(_h)
    logger.setLevel(logging.INFO)

# ---------------------------------------------------------------------------
# MS Graph OAuth scopes
# ---------------------------------------------------------------------------
_MS_GRAPH_SCOPES: List[str] = [
    "https://graph.microsoft.com/Mail.Read",
    "https://graph.microsoft.com/Mail.ReadWrite",
    "https://graph.microsoft.com/Files.ReadWrite",
    "https://graph.microsoft.com/User.Read",
]


# ---------------------------------------------------------------------------
# LangGraph state schema
# ---------------------------------------------------------------------------
class _EmailAgentState(TypedDict):
    emails:              List[dict]
    categories:          List[dict]
    user_filter:         str
    categorized_emails:  List[dict]
    filtered_emails:     List[dict]
    processed_files:     List[dict]
    errors:              List[str]
    # Runtime context (injected by OutlookAgentModule.run)
    _token:              str
    _processed_ids:      list
    _onedrive_folder:    str
    _openai_client:      Any
    _openai_model:       str
    _excel_log_path:     str
    _processed_ids_file: str


# ===========================================================================
# Internal helpers
# ===========================================================================

def _graph_get(token: str, url: str, params: Optional[dict] = None) -> dict:
    """Authenticated GET against the MS Graph API."""
    resp = _requests.get(url, headers={"Authorization": f"Bearer {token}"}, params=params)
    resp.raise_for_status()
    return resp.json()


def _mark_read(token: str, email_id: str) -> None:
    """Mark an Outlook message as read via Graph API."""
    try:
        _requests.patch(
            f"https://graph.microsoft.com/v1.0/me/messages/{email_id}",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json={"isRead": True},
        )
    except Exception as exc:
        logger.warning(f"Could not mark email as read: {exc}")


def _parse_filter(user_filter: str) -> List[str]:
    """Parse comma-separated filter string into lowercase category names."""
    s = user_filter.strip().lower()
    if s in ("all", "all categories", ""):
        return ["all"]
    return [p.strip() for p in s.split(",") if p.strip()]


def _extract_email(sender: str) -> str:
    """Extract bare email address from 'Name <addr>' or plain address."""
    if not sender:
        return ""
    m = re.search(r"<([^>]+)>", sender)
    if m:
        return m.group(1).strip()
    bare = sender.strip()
    return bare if "@" in bare else bare


def _safe_get_sender(msg: dict) -> str:
    """Safely extracts sender address from MS Graph message dictionary."""
    if not isinstance(msg, dict):
        return "Unknown"
    from_val = msg.get("from") or msg.get("sender")
    if isinstance(from_val, dict):
        email_addr = from_val.get("emailAddress")
        if isinstance(email_addr, dict):
            return str(email_addr.get("address", "Unknown"))
        elif isinstance(email_addr, str):
            return email_addr
        return str(from_val.get("address", "Unknown"))
    elif isinstance(from_val, str):
        return from_val
    return "Unknown"


def _safe_get_body(msg: dict) -> str:
    """Safely extracts body text from MS Graph message dictionary."""
    if not isinstance(msg, dict):
        return ""
    body_val = msg.get("body")
    if isinstance(body_val, dict):
        return str(body_val.get("content", ""))
    elif isinstance(body_val, str):
        return body_val
    return ""



# ===========================================================================
# Excel helpers  (self-contained; no dependency on excel_tracker.py)
# ===========================================================================
_EXCEL_SHEET   = "Processed Mails"
_EXCEL_HEADERS = ["Date", "Mail ID", "File Name", "Location"]
_HDR_FONT      = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HDR_FILL      = PatternFill(fill_type="solid", fgColor="2E75B6")
_HDR_ALIGN     = Alignment(horizontal="center", vertical="center")
_CELL_ALIGN    = Alignment(vertical="center", wrap_text=False)
_CELL_FONT     = Font(name="Calibri", size=10)
_THIN_BORDER   = Border(
    left=Side(style="thin",   color="D9D9D9"),
    right=Side(style="thin",  color="D9D9D9"),
    top=Side(style="thin",    color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
_COL_WIDTHS: Dict[str, int] = {
    "Date": 22, "Mail ID": 36, "File Name": 40, "Location": 70
}


def _excel_new_workbook() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = _EXCEL_SHEET
    for i, hdr in enumerate(_EXCEL_HEADERS, 1):
        cell = ws.cell(row=1, column=i, value=hdr)
        cell.font, cell.fill, cell.alignment, cell.border = (
            _HDR_FONT, _HDR_FILL, _HDR_ALIGN, _THIN_BORDER
        )
    ws.row_dimensions[1].height = 20
    for i, hdr in enumerate(_EXCEL_HEADERS, 1):
        ws.column_dimensions[get_column_letter(i)].width = _COL_WIDTHS.get(hdr, 20)
    ws.freeze_panes = "A2"
    return wb


def _excel_style_row(ws, row_idx: int) -> None:
    fill = PatternFill(
        fill_type="solid",
        fgColor="EBF3FB" if row_idx % 2 == 0 else "FFFFFF"
    )
    for col_idx in range(1, len(_EXCEL_HEADERS) + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if not cell.hyperlink:
            cell.font = _CELL_FONT
        cell.alignment, cell.fill, cell.border = _CELL_ALIGN, fill, _THIN_BORDER
    ws.row_dimensions[row_idx].height = 16


def _excel_append(records: List[Dict[str, Any]], excel_log_path: str) -> None:
    """Append processed-file records to the Excel log (creates file if absent)."""
    if not records:
        return
    os.makedirs(os.path.dirname(excel_log_path), exist_ok=True)
    if os.path.exists(excel_log_path):
        try:
            wb = openpyxl.load_workbook(excel_log_path)
            ws = (
                wb[_EXCEL_SHEET]
                if _EXCEL_SHEET in wb.sheetnames
                else wb.create_sheet(_EXCEL_SHEET)
            )
        except Exception:
            wb = _excel_new_workbook()
            ws = wb.active
    else:
        wb = _excel_new_workbook()
        ws = wb.active

    for rec in records:
        raw_ts = rec.get("processed_at", datetime.now())
        date_str = (
            raw_ts.strftime("%Y-%m-%d %H:%M:%S")
            if isinstance(raw_ts, datetime)
            else str(raw_ts)
        )
        nr = ws.max_row + 1
        ws.cell(row=nr, column=1, value=date_str)
        ws.cell(
            row=nr, column=2,
            value=_extract_email(rec.get("sender", "")) or rec.get("email_id", "")
        )
        ws.cell(row=nr, column=3, value=rec.get("filename", ""))
        loc = rec.get("path", "")
        loc_cell = ws.cell(row=nr, column=4, value=loc)
        if loc.startswith("http"):
            loc_cell.hyperlink = loc
            loc_cell.font = Font(
                name="Calibri", size=10, underline="single", color="0563C1"
            )
        _excel_style_row(ws, nr)

    for attempt in range(1, 6):
        try:
            wb.save(excel_log_path)
            logger.info(f"Excel log updated: {excel_log_path}")
            return
        except PermissionError:
            if attempt < 5:
                logger.warning(f"Excel locked (attempt {attempt}/5). Retrying in 3s...")
                time.sleep(3)
            else:
                logger.error("Could not save Excel log -- file still locked.")
        except Exception as exc:
            logger.error(f"Excel save error: {exc}")
            return


# ===========================================================================
# Processed-ID persistence  (deduplication across runs)
# ===========================================================================

def _load_processed_ids(path: str) -> Set[str]:
    dir_ = os.path.dirname(path)
    if dir_:
        os.makedirs(dir_, exist_ok=True)
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8-sig") as fh:
                return set(json.load(fh))
        except Exception:
            pass
    return set()


def _save_processed_ids(ids: Set[str], path: str) -> None:
    dir_ = os.path.dirname(path)
    if dir_:
        os.makedirs(dir_, exist_ok=True)
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(list(ids), fh, indent=2)


# ===========================================================================
# LangGraph node functions
# ===========================================================================

def _node_monitor_email(state: _EmailAgentState) -> dict:
    """Fetch new unread emails from Outlook via Microsoft Graph API."""
    logger.info("Node: monitor_email -- fetching unread Outlook emails...")
    errors        = list(state.get("errors", []))
    emails:       List[dict] = []
    token         = state["_token"]
    processed_ids = set(state.get("_processed_ids", []))

    try:
        data = _graph_get(
            token,
            "https://graph.microsoft.com/v1.0/me/messages",
            params={
                "$filter": "isRead eq false",
                "$top":    50,
                "$select": "id,subject,from,receivedDateTime,body,hasAttachments",
            },
        )
        messages = data.get("value", []) if isinstance(data, dict) else []
        logger.info(f"  Found {len(messages)} unread message(s).")

        for msg in messages:
            if not isinstance(msg, dict):
                continue
            msg_id = msg.get("id")
            if not msg_id or msg_id in processed_ids:
                continue

            attachments: List[dict] = []
            if msg.get("hasAttachments"):
                att_data = _graph_get(
                    token,
                    f"https://graph.microsoft.com/v1.0/me/messages/{msg_id}/attachments",
                )
                att_list = att_data.get("value", []) if isinstance(att_data, dict) else []
                for att in att_list:
                    if isinstance(att, dict) and att.get("@odata.type") == "#microsoft.graph.fileAttachment":
                        attachments.append({
                            "filename":      att.get("name", "attachment.pdf"),
                            "attachment_id": att.get("id", ""),
                            "message_id":    msg_id,
                            "content_bytes": att.get("contentBytes", ""),
                        })

            emails.append({
                "id":          msg_id,
                "subject":     msg.get("subject", "(No Subject)"),
                "sender":      _safe_get_sender(msg),
                "date":        msg.get("receivedDateTime", ""),
                "body":        _safe_get_body(msg)[:2000],
                "attachments": attachments,
            })

        logger.info(f"  {len(emails)} new email(s) to process.")
    except Exception as exc:
        err = f"monitor_email error: {exc}"
        logger.error(err)
        errors.append(err)

    return {"emails": emails, "errors": errors}


def _node_load_categories(state: _EmailAgentState) -> dict:
    """Passthrough: categories are pre-loaded into state by the runner."""
    logger.info("Node: load_categories -- categories already pre-loaded.")
    return {}


def _node_categorize_email(state: _EmailAgentState) -> dict:
    """Classify each email using OpenAI."""
    logger.info("Node: categorize_email -- classifying with OpenAI...")
    errors      = list(state.get("errors", []))
    categorized = []
    categories  = state.get("categories", [])
    oa_client   = state["_openai_client"]
    oa_model    = state["_openai_model"]
    user_filter = state.get("user_filter", "").strip()

    # "ALL" (uppercase) = download everything without per-category sorting
    if user_filter.upper() == "ALL":
        categorized = [{**e, "category": "ALL"} for e in state.get("emails", [])]
        logger.info("  Filter=ALL -- skipping OpenAI classification.")
        return {"categorized_emails": categorized, "errors": errors}

    if not categories:
        logger.warning("  No categories loaded -- all emails will be 'uncategorized'.")

    cat_list    = "\n".join(f"- {c['name']}: {c['description']}" for c in categories)
    valid_names = [c["name"] for c in categories]

    for email in state.get("emails", []):
        try:
            att_names = [
                a.get("filename", "")
                for a in email.get("attachments", [])
                if a.get("filename")
            ]
            resp = oa_client.chat.completions.create(
                model=oa_model,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            f"You are an email classifier. Classify the email into EXACTLY ONE "
                            f"category from this list:\n\n{cat_list}\n\n"
                            "Return ONLY the category name -- no punctuation, no explanation. "
                            "If none match, return: uncategorized"
                        ),
                    },
                    {
                        "role": "user",
                        "content": (
                            f"Subject: {email.get('subject', '')}\n"
                            f"Attachments: {', '.join(att_names) or 'None'}\n\n"
                            f"Body: {email.get('body', '')[:2000]}"
                        ),
                    },
                ],
                temperature=0.1,
                max_tokens=10,
            )
            # Universal Token Monitor
            try:
                import sys as _sys, os as _os
                _cp = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..')
                if _cp not in _sys.path: _sys.path.insert(0, _cp)
                from core.universal_token_monitor import track_usage as _tm
                _tm(resp.usage, model=oa_model, poc_name="outlook-agent",
                    file_name=email.get('subject', 'email')[:80], step_name="email_classification")
            except Exception: pass
            raw      = resp.choices[0].message.content.strip()
            category = next(
                (n for n in valid_names if n.lower() == raw.lower()), "uncategorized"
            )
            logger.info(f"  '{email.get('subject', '')[:60]}' -> {category}")
            categorized.append({**email, "category": category})
        except Exception as exc:
            err = f"Categorization error: {exc}"
            logger.error(err)
            errors.append(err)
            categorized.append({**email, "category": "uncategorized"})

    return {"categorized_emails": categorized, "errors": errors}


def _node_apply_filter(state: _EmailAgentState) -> dict:
    """Keep only emails whose category matches the user filter."""
    parsed = _parse_filter(state.get("user_filter", "all"))
    emails = state.get("categorized_emails", [])
    if "all" in parsed:
        filtered = emails
    else:
        filtered = [e for e in emails if e.get("category", "").lower() in parsed]
    logger.info(f"Node: apply_user_filter -- {len(filtered)}/{len(emails)} emails kept.")
    return {"filtered_emails": filtered}


def _node_download_and_store(state: _EmailAgentState) -> dict:
    """Upload attachments to OneDrive; persist tracking in Excel + processed_ids.json."""
    logger.info("Node: download_and_store -- uploading to OneDrive...")
    errors          = list(state.get("errors", []))
    processed_files = list(state.get("processed_files", []))
    token           = state["_token"]
    onedrive_folder = state["_onedrive_folder"]
    user_filter     = state.get("user_filter", "").strip()
    processed_ids   = set(state.get("_processed_ids", []))
    excel_log_path  = state["_excel_log_path"]
    id_file         = state["_processed_ids_file"]
    new_records:    List[dict] = []

    for email in state.get("filtered_emails", []):
        email_id    = email["id"]
        category    = email.get("category", "uncategorized")
        attachments = email.get("attachments", [])

        if not attachments:
            logger.info("  No attachments in email -- skipping.")
            _mark_read(token, email_id)
            processed_ids.add(email_id)
            continue

        target_name = "ALL files folder" if user_filter.upper() == "ALL" else category

        for att in attachments:
            try:
                filename   = att["filename"]
                file_bytes = base64.b64decode(att["content_bytes"])
                od_path    = f"{onedrive_folder}/{target_name}/{filename}"
                upload_url = (
                    f"https://graph.microsoft.com/v1.0/me/drive/root:/{od_path}:/content"
                )

                resp = _requests.put(
                    upload_url,
                    headers={
                        "Authorization": f"Bearer {token}",
                        "Content-Type":  "application/octet-stream",
                    },
                    data=file_bytes,
                )
                resp.raise_for_status()

                dest_link = resp.json().get("webUrl", od_path)
                logger.info(f"  Uploaded -> {dest_link}")

                record = {
                    "email_id":     email_id,
                    "sender":       email.get("sender", ""),
                    "category":     category,
                    "filename":     filename,
                    "path":         dest_link,
                    "size_bytes":   len(file_bytes),
                    "processed_at": datetime.now(),
                }
                processed_files.append(record)
                new_records.append(record)

            except Exception as exc:
                err = f"Upload error for '{att.get('filename', '?')}': {exc}"
                logger.error(err)
                errors.append(err)

        _mark_read(token, email_id)
        processed_ids.add(email_id)

    _save_processed_ids(processed_ids, id_file)
    if new_records and excel_log_path:
        _excel_append(new_records, excel_log_path)

    return {
        "processed_files": processed_files,
        "errors":          errors,
        "_processed_ids":  list(processed_ids),
    }


def _node_log_result(state: _EmailAgentState) -> dict:
    """Build and log the final processing summary."""
    categorized     = state.get("categorized_emails", [])
    processed_files = state.get("processed_files", [])
    errors          = state.get("errors", [])

    cat_counts: Dict[str, int] = {}
    for e in categorized:
        cat = e.get("category", "uncategorized")
        cat_counts[cat] = cat_counts.get(cat, 0) + 1

    files_by_cat: Dict[str, List[str]] = {}
    for f in processed_files:
        files_by_cat.setdefault(f.get("category", "uncategorized"), []).append(f["path"])

    summary = {
        "total_emails_processed": len(categorized),
        "categories_matched":     cat_counts,
        "files_uploaded":         len(processed_files),
        "files_by_category":      files_by_cat,
        "errors":                 errors,
    }
    logger.info("Summary:\n%s", json.dumps(summary, indent=2, default=str))
    return {"_summary": summary}


# ---------------------------------------------------------------------------
# Compile LangGraph workflow  (monitor -> load -> categorize -> filter -> store -> log)
# ---------------------------------------------------------------------------

def _build_workflow():
    wf = StateGraph(_EmailAgentState)
    wf.add_node("monitor_email",      _node_monitor_email)
    wf.add_node("load_categories",    _node_load_categories)
    wf.add_node("categorize_email",   _node_categorize_email)
    wf.add_node("apply_user_filter",  _node_apply_filter)
    wf.add_node("download_and_store", _node_download_and_store)
    wf.add_node("log_result",         _node_log_result)
    wf.add_edge(START,                "monitor_email")
    wf.add_edge("monitor_email",      "load_categories")
    wf.add_edge("load_categories",    "categorize_email")
    wf.add_edge("categorize_email",   "apply_user_filter")
    wf.add_edge("apply_user_filter",  "download_and_store")
    wf.add_edge("download_and_store", "log_result")
    wf.add_edge("log_result",         END)
    return wf.compile()


# ===========================================================================
# OutlookAgentModule -- public class
# ===========================================================================

class OutlookAgentModule:
    """
    Self-contained Outlook email attachment processing agent.

    All constructor parameters are optional; they fall back to environment
    variables loaded from .env.

    Parameters
    ----------
    azure_client_id     : Azure AD app (client) ID.
    azure_client_secret : App secret. If None, uses device-code flow.
    azure_tenant_id     : Azure AD tenant ID.
    openai_api_key      : OpenAI API key.
    openai_model        : Model for classification (default: gpt-4o).
    onedrive_folder     : Root OneDrive folder (default: AI_Agent_Attachments).
    categories          : List of {"name", "description"} dicts (overrides file).
    categories_file     : Path to categories.json.
    excel_log_path      : Path to Excel activity log.
    processed_ids_file  : Path to processed-IDs JSON (deduplication).
    log_excel           : Write Excel log after each run (default: True).
    token_cache_path    : Where to cache the MSAL device-code token.
    log_level           : Python logging level (default: INFO).
    """

    def __init__(
        self,
        azure_client_id:     Optional[str]        = None,
        azure_client_secret: Optional[str]        = None,
        azure_tenant_id:     Optional[str]        = None,
        openai_api_key:      Optional[str]        = None,
        openai_model:        str                  = "gpt-4o",
        onedrive_folder:     str                  = "AI_Agent_Attachments",
        categories:          Optional[List[dict]] = None,
        categories_file:     Optional[str]        = None,
        excel_log_path:      Optional[str]        = None,
        processed_ids_file:  Optional[str]        = None,
        log_excel:           bool                 = True,
        token_cache_path:    Optional[str]        = None,
        log_level:           int                  = logging.INFO,
    ) -> None:
        _base = os.path.dirname(os.path.abspath(__file__))
        _cfg  = os.path.join(_base, "config")
        _logs = os.path.join(_base, "logs")

        self.azure_client_id     = azure_client_id     or os.getenv("AZURE_CLIENT_ID",     "")
        self.azure_client_secret = azure_client_secret or os.getenv("AZURE_CLIENT_SECRET")
        self.azure_tenant_id     = azure_tenant_id     or os.getenv("AZURE_TENANT_ID",     "")
        self.onedrive_folder     = onedrive_folder     or os.getenv("ONEDRIVE_FOLDER",     "AI_Agent_Attachments")
        self.openai_api_key      = openai_api_key      or os.getenv("OPENAI_API_KEY",      "")
        self.openai_model        = openai_model
        self._openai_client      = OpenAI(api_key=self.openai_api_key)
        self.categories_file     = (
            categories_file    or os.getenv("CATEGORIES_FILE",    os.path.join(_cfg,  "categories.json"))
        )
        self.excel_log_path      = (
            excel_log_path     or os.getenv("EXCEL_LOG_PATH",     os.path.join(_logs, "processed_mails_log.xlsx"))
        )
        self.processed_ids_file  = (
            processed_ids_file or os.getenv("PROCESSED_IDS_FILE", os.path.join(_cfg,  "processed_ids.json"))
        )
        self.token_cache_path    = token_cache_path or os.path.join(_cfg, "ms_token_cache.json")
        self.log_excel           = log_excel
        self._categories         = categories
        logger.setLevel(log_level)
        self._workflow           = _build_workflow()

    # -----------------------------------------------------------------------
    # Authentication
    # -----------------------------------------------------------------------
    def get_access_token(self, refresh_token: Optional[str] = None) -> str:
        """
        Obtain a valid MS Graph access token.

        Priority:
          1. refresh_token supplied  -> ConfidentialClientApplication (silent)
          2. token cache on disk     -> PublicClientApplication (silent)
          3. Fallback                -> device-code interactive flow (one-time)
        """
        authority = f"https://login.microsoftonline.com/{self.azure_tenant_id}"

        # Refresh-token path (web-app OAuth, no UI required)
        if refresh_token:
            app    = msal.ConfidentialClientApplication(
                self.azure_client_id,
                authority=authority,
                client_credential=self.azure_client_secret,
            )
            result = app.acquire_token_by_refresh_token(
                refresh_token, scopes=_MS_GRAPH_SCOPES
            )
            if "access_token" not in result:
                raise RuntimeError(
                    f"Token refresh failed: {result.get('error_description')}"
                )
            logger.info("Token obtained via refresh_token.")
            return result["access_token"]

        # Device-code path (interactive once, then cached)
        cache = msal.SerializableTokenCache()
        if os.path.exists(self.token_cache_path):
            with open(self.token_cache_path) as fh:
                cache.deserialize(fh.read())

        app      = msal.PublicClientApplication(
            self.azure_client_id, authority=authority, token_cache=cache
        )
        accounts = app.get_accounts()
        result   = None
        if accounts:
            result = app.acquire_token_silent(_MS_GRAPH_SCOPES, account=accounts[0])

        if not result:
            logger.info("No cached token; initiating device-code flow...")
            flow   = app.initiate_device_flow(scopes=_MS_GRAPH_SCOPES)
            print(f"\n{flow['message']}\n")
            result = app.acquire_token_by_device_flow(flow)

        if cache.has_state_changed:
            os.makedirs(os.path.dirname(self.token_cache_path), exist_ok=True)
            with open(self.token_cache_path, "w") as fh:
                fh.write(cache.serialize())

        if "access_token" not in result:
            raise RuntimeError(
                f"MSAL auth failed: {result.get('error_description')}"
            )
        logger.info("Token obtained (MSAL cache / device-code).")
        return result["access_token"]

    # -----------------------------------------------------------------------
    # Categories
    # -----------------------------------------------------------------------
    def load_categories(self) -> List[dict]:
        """Return categories from constructor, or load from categories.json."""
        if self._categories is not None:
            return self._categories
        try:
            with open(self.categories_file, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            cats = data.get("categories", [])
            logger.info(f"Loaded {len(cats)} categories from {self.categories_file}")
            return cats
        except FileNotFoundError:
            logger.warning(f"categories.json not found: {self.categories_file}")
            return []
        except json.JSONDecodeError as exc:
            logger.error(f"Invalid JSON in categories.json: {exc}")
            return []

    def add_category(self, name: str, description: str) -> bool:
        """
        Add a new category to categories.json (persistent across runs).
        Returns True if added, False if already existed.
        """
        try:
            with open(self.categories_file, "r", encoding="utf-8") as fh:
                data = json.load(fh)
        except (FileNotFoundError, json.JSONDecodeError):
            data = {"categories": []}

        if name.lower() in [c["name"].lower() for c in data.get("categories", [])]:
            logger.info(f"Category '{name}' already exists.")
            return False

        data.setdefault("categories", []).append(
            {"name": name, "description": description}
        )
        os.makedirs(os.path.dirname(self.categories_file), exist_ok=True)
        with open(self.categories_file, "w", encoding="utf-8") as fh:
            json.dump(data, fh, indent=2)
        self._categories = None   # invalidate in-memory cache
        logger.info(f"Category '{name}' added to {self.categories_file}.")
        return True

    # -----------------------------------------------------------------------
    # Core run method
    # -----------------------------------------------------------------------
    def run(
        self,
        user_filter:   str                      = "all",
        refresh_token: Optional[str]            = None,
        custom_paths:  Optional[Dict[str, str]] = None,
    ) -> Dict[str, Any]:
        """
        Execute one full pipeline pass.

        Parameters
        ----------
        user_filter   : "all"            -> sort every email into category folders
                        "ALL"            -> dump everything into one folder
                        "Invoice, Acord" -> comma-separated category names
        refresh_token : OAuth refresh token (avoids device-code prompt).
        custom_paths  : {category_lower: onedrive_subfolder} name overrides.

        Returns
        -------
        dict with keys:
            total_emails_processed  int
            categories_matched      dict[str, int]
            files_uploaded          int
            files_by_category       dict[str, list[str]]
            errors                  list[str]
        """
        logger.info(f"OutlookAgentModule.run() -- filter={user_filter!r}")

        token         = self.get_access_token(refresh_token=refresh_token)
        categories    = self.load_categories()
        processed_ids = _load_processed_ids(self.processed_ids_file)

        initial: _EmailAgentState = {
            "emails":              [],
            "categories":          categories,
            "user_filter":         user_filter,
            "categorized_emails":  [],
            "filtered_emails":     [],
            "processed_files":     [],
            "errors":              [],
            "_token":              token,
            "_processed_ids":      list(processed_ids),
            "_onedrive_folder":    self.onedrive_folder,
            "_openai_client":      self._openai_client,
            "_openai_model":       self.openai_model,
            "_excel_log_path":     self.excel_log_path if self.log_excel else "",
            "_processed_ids_file": self.processed_ids_file,
        }

        try:
            final = self._workflow.invoke(initial)
        except Exception as exc:
            logger.error(f"Workflow error: {exc}")
            return {"errors": [str(exc)]}

        return final.get("_summary", {
            "total_emails_processed": 0,
            "categories_matched":     {},
            "files_uploaded":         0,
            "files_by_category":      {},
            "errors":                 final.get("errors", []),
        })

    # -----------------------------------------------------------------------
    # Continuous polling
    # -----------------------------------------------------------------------
    def run_loop(
        self,
        user_filter:      str                      = "all",
        interval_seconds: int                      = 30,
        refresh_token:    Optional[str]            = None,
        custom_paths:     Optional[Dict[str, str]] = None,
        max_iterations:   Optional[int]            = None,
    ) -> None:
        """
        Poll Outlook continuously, calling run() every interval_seconds.

        Parameters
        ----------
        user_filter      : Same as run().
        interval_seconds : Seconds between polls (default: 30).
        refresh_token    : OAuth refresh token.
        custom_paths     : Path overrides.
        max_iterations   : Stop after N iterations (None = run forever).
                           Useful for automated testing.
        """
        logger.info(
            f"Continuous monitoring started -- filter={user_filter!r}, "
            f"interval={interval_seconds}s. Press Ctrl+C to stop."
        )
        iteration = 0
        while True:
            if max_iterations is not None and iteration >= max_iterations:
                logger.info(f"Reached max_iterations={max_iterations}. Stopping.")
                break

            summary = self.run(
                user_filter=user_filter,
                refresh_token=refresh_token,
                custom_paths=custom_paths,
            )
            iteration += 1
            logger.info(
                f"Iteration {iteration} complete -- "
                f"{summary.get('files_uploaded', 0)} file(s) uploaded. "
                f"Waiting {interval_seconds}s..."
            )
            try:
                time.sleep(interval_seconds)
            except KeyboardInterrupt:
                logger.info("Stopped by user (Ctrl+C).")
                break

    # -----------------------------------------------------------------------
    # Convenience: classify one email (no full pipeline)
    # -----------------------------------------------------------------------
    def classify_email(
        self,
        subject:          str,
        body:             str,
        attachment_names: Optional[List[str]] = None,
    ) -> str:
        """
        Classify a single email using OpenAI without running the full pipeline.
        Returns the matched category name or 'uncategorized'.
        """
        categories = self.load_categories()
        cat_list   = "\n".join(
            f"- {c['name']}: {c['description']}" for c in categories
        )
        valid      = [c["name"] for c in categories]
        att_str    = ", ".join(attachment_names or []) or "None"

        resp = self._openai_client.chat.completions.create(
            model=self.openai_model,
            messages=[
                {
                    "role": "system",
                    "content": (
                        f"Classify the email into EXACTLY ONE category:\n\n{cat_list}\n\n"
                        "Return ONLY the category name. "
                        "If none match, return: uncategorized"
                    ),
                },
                {
                    "role": "user",
                    "content": (
                        f"Subject: {subject}\n"
                        f"Attachments: {att_str}\n\n"
                        f"Body: {body[:2000]}"
                    ),
                },
            ],
            temperature=0.1,
            max_tokens=10,
        )
        # Universal Token Monitor
        try:
            import sys as _sys, os as _os
            _cp = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..')
            if _cp not in _sys.path: _sys.path.insert(0, _cp)
            from core.universal_token_monitor import track_usage as _tm
            _tm(resp.usage, model=self.openai_model, poc_name="outlook-agent",
                file_name=subject[:80], step_name="standalone_email_classification")
        except Exception: pass
        raw = resp.choices[0].message.content.strip()
        return next((n for n in valid if n.lower() == raw.lower()), "uncategorized")

    # -----------------------------------------------------------------------
    # Convenience: fetch raw emails only
    # -----------------------------------------------------------------------
    def fetch_unread_emails(
        self,
        refresh_token:  Optional[str] = None,
        limit:          int           = 50,
        skip_processed: bool          = True,
    ) -> List[dict]:
        """
        Fetch unread Outlook emails without classification or upload.

        Returns list of dicts:
            {id, subject, sender, date, body, attachments}
        """
        token         = self.get_access_token(refresh_token=refresh_token)
        processed_ids = (
            _load_processed_ids(self.processed_ids_file) if skip_processed else set()
        )
        emails = []

        data = _graph_get(
            token,
            "https://graph.microsoft.com/v1.0/me/messages",
            params={
                "$filter": "isRead eq false",
                "$top":    limit,
                "$select": "id,subject,from,receivedDateTime,body,hasAttachments",
            },
        )
        messages = data.get("value", []) if isinstance(data, dict) else []
        for msg in messages:
            if not isinstance(msg, dict):
                continue
            msg_id = msg.get("id")
            if not msg_id or msg_id in processed_ids:
                continue
            attachments = []
            if msg.get("hasAttachments"):
                att_data = _graph_get(
                    token,
                    f"https://graph.microsoft.com/v1.0/me/messages/{msg_id}/attachments",
                )
                att_list = att_data.get("value", []) if isinstance(att_data, dict) else []
                for att in att_list:
                    if isinstance(att, dict) and att.get("@odata.type") == "#microsoft.graph.fileAttachment":
                        attachments.append({
                            "filename":      att.get("name", "attachment.pdf"),
                            "attachment_id": att.get("id", ""),
                            "message_id":    msg_id,
                            "content_bytes": att.get("contentBytes", ""),
                        })
            emails.append({
                "id":          msg_id,
                "subject":     msg.get("subject", "(No Subject)"),
                "sender":      _safe_get_sender(msg),
                "date":        msg.get("receivedDateTime", ""),
                "body":        _safe_get_body(msg)[:2000],
                "attachments": attachments,
            })

        logger.info(f"fetch_unread_emails: {len(emails)} email(s) returned.")
        return emails

    def __repr__(self) -> str:
        return (
            f"OutlookAgentModule("
            f"tenant={self.azure_tenant_id[:8]}..., "
            f"model={self.openai_model}, "
            f"onedrive={self.onedrive_folder!r})"
        )


# ===========================================================================
# Standalone entry point
#   python outlook_agent_module.py [--filter X] [--loop] [--interval N]
# ===========================================================================
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Outlook Agent Module -- standalone runner"
    )
    parser.add_argument("--filter",     default="all",
                        help="Category filter (default: all)")
    parser.add_argument("--loop",       action="store_true",
                        help="Run in continuous polling mode")
    parser.add_argument("--interval",   type=int, default=30,
                        help="Poll interval in seconds (default: 30)")
    parser.add_argument("--iterations", type=int, default=None,
                        help="Max loop iterations for testing (default: unlimited)")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.FileHandler("agent.log", encoding="utf-8"),
            logging.StreamHandler(),
        ],
    )

    agent = OutlookAgentModule()   # reads credentials from .env

    if args.loop:
        agent.run_loop(
            user_filter=args.filter,
            interval_seconds=args.interval,
            max_iterations=args.iterations,
        )
    else:
        summary = agent.run(user_filter=args.filter)
        print("\n" + "=" * 60)
        print("RESULT")
        print("=" * 60)
        print(json.dumps(summary, indent=2, default=str))
