"""
Invoice-to-Census Renewal Premium Pipeline
===========================================
Extracts plan-level renewal premium rates from a Benefits Renewal PDF (image-based)
using OpenAI GPT-4o vision, then matches each census employee's current plan + coverage
tier to look up the correct renewal premium and writes it into the census.

Phases:
  1 – Ingest census spreadsheet (handles merged cells, header offset)
  2 – OCR / Vision-extract renewal rate table from PDF invoice
  3 – Match census rows to renewal rates by (plan_name, coverage_tier)
  4 – Write renewal amounts into census output
  5 – Generate audit log
  6 – Save extracted text to extracted_text folder

Usage:
    python invoice_census_audit.py ^
        --census   path/to/census.xlsx ^
        --invoices path/to/invoice.pdf  ^
        --out-census path/to/census_updated.xlsx ^
        --out-log    path/to/audit_log.xlsx ^
        [--census-sheet "Census"] ^
        [--header-row 21] ^
        [--use-llm]
"""

import argparse
import json
import logging
import os
import re
import sys
import base64
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from dotenv import load_dotenv

# ──────────────────────────────────────────────
# Load environment variables
# ──────────────────────────────────────────────
load_dotenv()

# ──────────────────────────────────────────────
# Logging
# ──────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ──────────────────────────────────────────────
# Data structures
# ──────────────────────────────────────────────
@dataclass
class RenewalRate:
    """A single plan + tier renewal rate extracted from the invoice PDF."""
    plan_name: str          # e.g. "AETNA-OA EPO 1500-80%"
    tier: str               # e.g. "Employee", "Employee + Spouse", "Employee + Children", "Employee + Family"
    current_monthly: float  # Current monthly premium
    renewal_monthly: float  # Renewal monthly premium
    source_page: int = 0


@dataclass
class CensusRow:
    """A single row from the census spreadsheet."""
    excel_row: int             # 1-based Excel row number
    ee_row: Optional[int]      # EE Row number from column A
    first_name: str
    last_name: str
    gender: str
    dob: str
    relationship: str          # EE, SP, CH
    dependent_of: Optional[int]  # Dependent of Employee Row
    coverage_type: str         # EE, FAM, WO, etc.
    cobra: str
    current_plan: Optional[str]      # Current Plan Description (col K)
    monthly_premium: Optional[float]  # Monthly Total Premium (col L)
    renewal_amount: Optional[float]   # Renewal (col M) - to be filled


@dataclass
class AuditRecord:
    excel_row: int
    ee_row: Optional[int]
    employee_name: str
    current_plan: Optional[str]
    coverage_type: str
    relationship: str
    current_premium: Optional[float]
    matched_renewal: Optional[float]
    match_method: str          # EXACT_PLAN_TIER | PLAN_ONLY | NO_MATCH | DEPENDENT | WAIVED
    status: str                # MATCH | NO_PLAN | WAIVED | DEPENDENT | NOT_FOUND
    notes: str = ""


# ──────────────────────────────────────────────
# Coverage tier mapping
# ──────────────────────────────────────────────
def coverage_to_tier(coverage_type: str, relationship: str) -> str:
    """
    Map census coverage type + relationship to invoice tier label.
    Coverage types: EE (employee only), FAM (family), EE+SP, EE+CH, WO (waived/opt-out)
    """
    cov = (coverage_type or "").upper().strip()
    rel = (relationship or "").upper().strip()

    if cov == "WO" or cov == "W/O" or cov == "WAIVED":
        return "WAIVED"

    # Dependents (SP, CH) don't get their own premium row
    if rel in ("SP", "CH", "SPOUSE", "CHILD"):
        return "DEPENDENT"

    if cov == "EE" or cov == "EMPLOYEE":
        return "Employee"
    elif cov == "FAM" or cov == "FAMILY" or cov == "EE+FAM":
        return "Employee + Family"
    elif cov in ("EE+SP", "ES", "EMP+SP"):
        return "Employee + Spouse"
    elif cov in ("EE+CH", "EC", "EMP+CH"):
        return "Employee + Children"
    else:
        # Default: if employee (EE relationship) with unknown coverage, treat as Employee
        if rel == "EE" or rel == "EMPLOYEE":
            return "Employee"
        return "UNKNOWN"


def normalize_plan_name(plan: str) -> str:
    """Normalize plan name for fuzzy matching."""
    if not plan:
        return ""
    p = plan.upper().strip()
    # Remove punctuation like commas
    p = p.replace(",", "")
    # Remove extra spaces
    p = re.sub(r"\s+", " ", p)
    # Standardize common patterns
    p = p.replace("AETNA-", "AETNA ")
    p = p.replace("CIGNA-", "CIGNA ")
    if p.startswith("CIGNA "):
        p = p[6:]
    if p.startswith("AETNA "):
        p = p[6:]
    p = p.replace("  ", " ")
    return p.strip()


# ──────────────────────────────────────────────
# Phase 1 – Census template detection
# ──────────────────────────────────────────────
def detect_census_type(ws, max_check_row: int = 5) -> int:
    """
    Auto-detect census template type by scanning the first few rows.

    Type 2 signatures: "Employee Name" + "Coverage Tier" + "Home Zip Code" must
    appear together with at least one of "Plan Enrolled" or "Plan Name" in one row.
    Returns 2 if Type 2, else 1 (default / Type 1).
    """
    type2_required = {"Employee Name", "Coverage Tier", "Home Zip Code"}
    type2_plan_variants = {"Plan Enrolled", "Plan Name"}
    for r in range(1, max_check_row + 1):
        row_vals = {
            str(ws.cell(row=r, column=c).value or "").strip()
            for c in range(1, ws.max_column + 1)
        }
        if type2_required.issubset(row_vals) and row_vals & type2_plan_variants:
            log.info("  Detected Type 2 census template at row %d", r)
            return 2
    log.info("  Detected Type 1 census template")
    return 1


# ──────────────────────────────────────────────
# Phase 1 – Ingest census (handles complex Excel layouts)
# ──────────────────────────────────────────────
def ingest_census(
    path: Path,
    sheet: str,
    header_row: int = 21,
) -> List[CensusRow]:
    """
    Read census from Excel with openpyxl directly (to handle merged cells).
    header_row is 1-based Excel row where headers are.
    Data starts at header_row + 1.
    """
    log.info("Phase 1 – Ingesting census: %s [sheet=%s, header_row=%d]", path, sheet, header_row)

    wb = load_workbook(str(path), data_only=True)
    ws = wb[sheet]

    # Read headers from the header row
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[str(val).strip()] = col

    log.info("  Headers found: %s", list(headers.keys()))

    # Map header names to column indices
    col_ee_row = headers.get("EE Row", headers.get("EE\nRow", 1))
    col_first_name = headers.get("First Name", 2)
    col_last_name = headers.get("Last Name", 3)
    col_gender = headers.get("Gender", 4)
    col_dob = headers.get("Date of Birth", 5)
    col_relationship = headers.get("Relation-ship to Employee",
                       headers.get("Relationship to Employee",
                       headers.get("Relationship", 7)))
    col_dependent_of = headers.get("Dependent of Employee Row",
                       headers.get("Dependent of\nEmployee Row", 8))
    col_coverage = headers.get("Coverage Type",
                   headers.get("Coverage\nType", 9))
    col_cobra = headers.get("COBRA", 10)
    col_plan = headers.get("Current Plan Description",
               headers.get("Current Plan\nDescription", 11))
    col_premium = headers.get("Monthly Total Premium",
                  headers.get("Monthly Total\nPremium", 12))
    col_renewal = headers.get("Renewal", 13)

    log.info("  Column mapping: Plan=%s, Premium=%s, Renewal=%s, Relationship=%s, Coverage=%s",
             col_plan, col_premium, col_renewal, col_relationship, col_coverage)

    rows: List[CensusRow] = []
    data_start = header_row + 1

    for r in range(data_start, ws.max_row + 1):
        # Check if row has any data
        first_name = ws.cell(row=r, column=col_first_name).value
        if not first_name:
            continue  # Skip empty rows

        ee_row_val = ws.cell(row=r, column=col_ee_row).value
        try:
            ee_row = int(ee_row_val) if ee_row_val is not None else None
        except (ValueError, TypeError):
            ee_row = None  # non-numeric value in EE Row column — treat as absent

        last_name = str(ws.cell(row=r, column=col_last_name).value or "")
        gender = str(ws.cell(row=r, column=col_gender).value or "")
        dob = str(ws.cell(row=r, column=col_dob).value or "")
        relationship = str(ws.cell(row=r, column=col_relationship).value or "")
        dep_of_val = ws.cell(row=r, column=col_dependent_of).value
        try:
            dependent_of = int(dep_of_val) if dep_of_val is not None else None
        except (ValueError, TypeError):
            dependent_of = None  # non-numeric value in Dependent Of column — treat as absent
        coverage = str(ws.cell(row=r, column=col_coverage).value or "")
        cobra = str(ws.cell(row=r, column=col_cobra).value or "")

        plan_val = ws.cell(row=r, column=col_plan).value
        plan = str(plan_val).strip() if plan_val else None

        premium_val = ws.cell(row=r, column=col_premium).value
        premium = None
        if premium_val is not None:
            try:
                premium = float(premium_val)
            except (ValueError, TypeError):
                pass

        renewal_val = ws.cell(row=r, column=col_renewal).value
        renewal = None
        if renewal_val is not None:
            try:
                renewal = float(renewal_val)
            except (ValueError, TypeError):
                pass

        rows.append(CensusRow(
            excel_row=r,
            ee_row=ee_row,
            first_name=str(first_name),
            last_name=last_name,
            gender=gender,
            dob=dob,
            relationship=relationship,
            dependent_of=dependent_of,
            coverage_type=coverage,
            cobra=cobra,
            current_plan=plan,
            monthly_premium=premium,
            renewal_amount=renewal,
        ))

    wb.close()
    log.info("  Loaded %d census rows", len(rows))
    return rows


def ingest_census_type2(
    path: Path,
    sheet: str,
    header_row: int = 1,
) -> List[CensusRow]:
    """
    Read a Type 2 census from Excel.

    Type 2 layout (flat, no dependent rows):
      Row 1 = headers, data starts at row 2.
      Columns: Employee Name | Gender | Date of Birth | Home Zip Code |
               Coverage Tier | COBRA | Plan Enrolled | Current Premium |
               Renewal | Notes

    Every row represents a standalone employee – no dependent rows,
    no EE-Row numbering.
    """
    log.info(
        "Phase 1 – Ingesting Type 2 census: %s [sheet=%s, header_row=%d]",
        path, sheet, header_row,
    )

    wb = load_workbook(str(path), data_only=True)
    ws = wb[sheet]

    # Read headers from the header row
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val:
            headers[str(val).strip()] = col

    log.info("  Headers found: %s", list(headers.keys()))

    # Column indices – use sensible defaults if a header is missing.
    # Accept alternative header names used by different census variants.
    col_name     = headers.get("Employee Name", 1)
    col_gender   = headers.get("Gender", 2)
    col_dob      = headers.get("Date of Birth", 3)
    col_coverage = headers.get("Coverage Tier", 5)
    col_cobra    = headers.get("COBRA", 6)
    col_plan     = headers.get("Plan Enrolled",
                   headers.get("Plan Name", 7))          # «Plan Name» variant
    col_premium  = headers.get("Current Premium",
                   headers.get("Monthly Premium", 8))    # «Monthly Premium» variant
    col_renewal  = headers.get("Renewal", 9)

    log.info(
        "  Column mapping: Name=%s, Coverage=%s, Plan=%s, Premium=%s, Renewal=%s",
        col_name, col_coverage, col_plan, col_premium, col_renewal,
    )

    rows: List[CensusRow] = []
    data_start = header_row + 1

    for r in range(data_start, ws.max_row + 1):
        name_val = ws.cell(row=r, column=col_name).value
        if not name_val:
            continue  # Skip empty rows

        full_name = str(name_val).strip()
        # Split into first / last on the first space; keep remainder as last name
        parts = full_name.split(" ", 1)
        first_name = parts[0]
        last_name  = parts[1] if len(parts) > 1 else ""

        gender   = str(ws.cell(row=r, column=col_gender).value or "")
        dob      = str(ws.cell(row=r, column=col_dob).value or "")
        coverage = str(ws.cell(row=r, column=col_coverage).value or "")
        cobra    = str(ws.cell(row=r, column=col_cobra).value or "")

        plan_val = ws.cell(row=r, column=col_plan).value
        plan = str(plan_val).strip() if plan_val else None

        premium_val = ws.cell(row=r, column=col_premium).value
        premium = None
        if premium_val is not None:
            try:
                premium = float(premium_val)
            except (ValueError, TypeError):
                pass

        renewal_val = ws.cell(row=r, column=col_renewal).value
        renewal = None
        if renewal_val is not None:
            try:
                renewal = float(renewal_val)
            except (ValueError, TypeError):
                pass

        rows.append(CensusRow(
            excel_row=r,
            ee_row=None,          # Type 2 has no EE-Row numbering
            first_name=first_name,
            last_name=last_name,
            gender=gender,
            dob=dob,
            relationship="EE",    # Every Type 2 row is an employee
            dependent_of=None,
            coverage_type=coverage,
            cobra=cobra,
            current_plan=plan,
            monthly_premium=premium,
            renewal_amount=renewal,
        ))

    wb.close()
    log.info("  Loaded %d census rows (Type 2)", len(rows))
    return rows


# ──────────────────────────────────────────────
# Phase 2 – Extract renewal rates from invoice PDF
# ──────────────────────────────────────────────
_doctr_predictor = None

def get_doctr_predictor():
    """Lazy-load and cache the DocTR OCR predictor."""
    global _doctr_predictor
    if _doctr_predictor is None:
        from doctr.models import ocr_predictor
        log.info("Initializing DocTR OCR Predictor (downloading model weights if first run)...")
        _doctr_predictor = ocr_predictor(pretrained=True)
    return _doctr_predictor


def extract_layout_preserving_text(pdf_path: Path, page_index: int, image) -> Tuple[str, str]:
    """
    Extract text from a page using the fallback hierarchy:
    1. DocTR (deep learning layout-preserving OCR)
    2. pdfplumber (reliable digital-PDF text extraction)
    3. Tesseract OCR (simple line-by-line OCR fallback)
    """
    # 1. Try DocTR
    try:
        from doctr.io import DocumentFile
        import tempfile
        import os
        
        log.info("  Page %d: Attempting DocTR layout-preserving OCR...", page_index)
        predictor = get_doctr_predictor()
        
        # Save PIL image to a temporary file to guarantee compatible path format for doctr
        with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
            image.save(tmp.name, format="PNG")
            tmp_path = tmp.name

        try:
            doc = DocumentFile.from_images([tmp_path])
            result = predictor(doc)
            page_text = result.render()
        finally:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

        if page_text and len(page_text.strip()) > 50:
            log.info("  Page %d: Extracted via DocTR (%d chars)", page_index, len(page_text))
            return page_text, "doctr"
    except Exception as e:
        log.warning("  Page %d: DocTR OCR failed, falling back: %s", page_index, e)

    # 2. Try pdfplumber
    try:
        import pdfplumber
        with pdfplumber.open(str(pdf_path)) as pdf:
            if page_index <= len(pdf.pages):
                page = pdf.pages[page_index - 1]
                page_text = page.extract_text()
                if page_text and len(page_text.strip()) > 50:
                    log.info("  Page %d: Extracted via pdfplumber (%d chars)", page_index, len(page_text))
                    return page_text, "pdfplumber"
    except Exception as e:
        log.warning("  Page %d: pdfplumber extraction failed: %s", page_index, e)

    # 3. Try Tesseract OCR
    try:
        import pytesseract
        text = pytesseract.image_to_string(image, config="--dpi 300 --psm 6")
        log.info("  Page %d: Extracted via Tesseract OCR (%d chars)", page_index, len(text))
        return text, "pytesseract"
    except Exception as e:
        log.error("  Page %d: Tesseract OCR failed: %s", page_index, e)
        return "", "failed"


def extract_pdf_text_page_by_page(pdf_path: Path, images: List) -> Tuple[str, Dict[int, str]]:
    """
    Extract text page by page from the PDF using the layout-preserving fallback chain.
    Returns (full_text, page_methods).
    """
    full_text = ""
    page_methods = {}
    for i, img in enumerate(images, 1):
        text, method = extract_layout_preserving_text(pdf_path, i, img)
        page_methods[i] = method
        full_text += f"\n=== PAGE {i} ===\n{text}\n"
    return full_text, page_methods


def extract_rates_with_ocr(pdf_path: Path, poppler_path: str = None) -> Tuple[List[RenewalRate], str]:
    """
    Extract renewal rates using OCR (Tesseract) + regex parsing.
    Returns (list of rates, full extracted text).
    """
    from pdf2image import convert_from_path

    log.info("Phase 2 – Extracting rates via OCR from: %s", pdf_path)

    kwargs = {"dpi": 300}
    if poppler_path:
        kwargs["poppler_path"] = poppler_path

    images = convert_from_path(str(pdf_path), **kwargs)
    all_rates: List[RenewalRate] = []

    full_text, _ = extract_pdf_text_page_by_page(pdf_path, images)

    log.info("  Text extraction completed for %d pages", len(images))
    return all_rates, full_text


def is_rate_page(text: str) -> bool:
    """Check if page text is likely to contain rate tables or enrollment details."""
    t = text.upper()
    
    # Exclude pages that are clearly not rate pages
    if "REQUIRED UNIFORM MODIFICATION NOTICE" in t:
        return False
    if "HOW TO RENEW" in t or "2 EASY WAYS TO RENEW" in t:
        return False
    if "TAX FORM SUBMISSION" in t or "TAX INSTRUCTIONS" in t:
        return False
    if "MEDICARE ESTIMATION" in t:
        return False
    if "AUTHORIZATION FOR BROKER TO ACT" in t:
        return False
    if "SMALL GROUP APPLICATION AND ANNUAL CERTIFICATION" in t:
        return False
    if "CLIENT BENEFIT ELECTION" in t and "ACKNOWLEDGEMENT" in t:
        return False
    if "HAWAII PREPAID HEALTH CARE ACT" in t:
        return False
    if "NEW YORK INSURANCE" in t:
        return False
    if "TEXAS SB51" in t:
        return False
    if "DISABILITY" in t:
        # Only skip if the page does not contain medical/dental rate tiers
        has_medical_tiers = any(ind in t for ind in ["+ SPOUSE", "+SPOUSE", "+ FAMILY", "+FAMILY", "+ CHILD", "+CHILD", "+ CHILDREN", "+CHILDREN", "EMPLOYEE +"])
        if not has_medical_tiers:
            return False
        
    # Rate indicators
    indicators = [
        "EMPLOYEE ONLY", "EMPLOYEEONLY",
        "RATES AND PREMIUMS", "CURRENTRATES", "RENEWALRATES",
        "MONTHLYRATE", "ENROLLMENTRATES",
        "APPENDIX A: EMPLOYEE", "EMPLOYEE ENROLLMENT DETAIL",
        "+ SPOUSE", "+SPOUSE",
        "+ CHILD", "+CHILD",
        "+ FAMILY", "+FAMILY",
        "+ CHILDREN", "+CHILDREN",
        "MONTHLY PREMIUM", "MONTHLY PREMIUMS",
        "MONTHLY RATE", "MONTHLY RATES",
        # Per-subscriber comparison grid (e.g. Anthem "Monthly Premium Comparison Details")
        "CURRENT RATE", "NEW RATE", "NEW MONTHLY", "PREMIUM COMPARISON",
        "INSURED SUBSCRIBERS", "CURRENT MONTHLY PREMIUM",
        "PLAN/CONTRIBUTION REPORT", "PLAN DETAIL REPORT",
        "PLAN/CONTRIBUTION", "PLAN DETAIL",
    ]
    
    for ind in indicators:
        if ind in t:
            return True
            
    # Check for at least two whole-word tier abbreviations: EE, ES, EC, EF
    tiers_found = sum(1 for tier in [r'\bEE\b', r'\bES\b', r'\bEC\b', r'\bEF\b'] if re.search(tier, t))
    if tiers_found >= 2:
        return True

    return False


def extract_rates_with_llm(pdf_path: Path, poppler_path: str = None) -> Tuple[List[RenewalRate], str]:
    """
    Extract renewal rates using OpenAI GPT-4o API (text-only).
    Sends the layout-preserved page text of relevant pages and asks for structured rate extraction.
    Returns (list of rates, full extracted text).
    """
    from pdf2image import convert_from_path
    from openai import OpenAI

    api_key = os.getenv("OPENAI_API_KEY")
    if not api_key:
        log.error("OPENAI_API_KEY not found in environment. Cannot use LLM extraction.")
        sys.exit(1)

    client = OpenAI(api_key=api_key)

    log.info("Phase 2 – Extracting rates via LLM (GPT-4o text-only) from: %s", pdf_path)

    kwargs = {"dpi": 300}
    if poppler_path:
        kwargs["poppler_path"] = poppler_path

    images = convert_from_path(str(pdf_path), **kwargs)
    all_rates: List[RenewalRate] = []

    # Do text extraction page-by-page using the layout-preserving fallback chain
    full_text, page_methods = extract_pdf_text_page_by_page(pdf_path, images)

    # Focus on pages that likely contain premium rate tables
    premium_pages = []
    for i in range(1, len(images) + 1):
        # Extract page text from full_text
        page_marker = f"=== PAGE {i} ==="
        next_marker = f"=== PAGE {i + 1} ==="
        page_text = ""
        if page_marker in full_text:
            try:
                start_idx = full_text.index(page_marker) + len(page_marker)
                if next_marker in full_text:
                    end_idx = full_text.index(next_marker)
                else:
                    end_idx = len(full_text)
                page_text = full_text[start_idx:end_idx].strip()
            except ValueError:
                pass
                
        if not is_rate_page(page_text):
            log.info("  Page %d: Skipping rate extraction (no rate tables/enrollment details found)", i)
            continue
            
        method = page_methods.get(i, "unknown")
        log.info("  Page %d: Selected for rate extraction (text extraction method: %s)", i, method)
        premium_pages.append((i, page_text))

    # Process pages with an overlapping window so that plans split across consecutive pages
    # (e.g. Current on page N, Renewal on page N+1) are always seen in the same LLM call.
    # We use a stride of 2 with a window of 3: batches are [0-2], [2-4], [4-6] ...
    # This means every boundary page is included in two consecutive batches, ensuring
    # cross-page continuations are always sent together.
    batch_stride = 2  # advance by 2 so each boundary page overlaps
    for batch_start in range(0, len(premium_pages), batch_stride):
        batch = premium_pages[batch_start:batch_start + 3]
        # Avoid sending duplicate-only batches at the tail
        if batch_start > 0 and len(batch) == 1 and len(premium_pages) % batch_stride == 1:
            break

        prompt_text = """Analyze these benefits renewal PDF pages. Extract ALL plan premium rate tables you find.

IMPORTANT WARNING ABOUT PLAN FEATURE / BENEFIT COMPARISON PAGES:
- A page containing plan features (such as deductibles, out-of-pocket maximums, coinsurance, office visit copays, prescription tiers, hospital care details, network lists) is NOT a premium rate table.
- DO NOT extract rates from these plan design / feature comparison pages. If a page only lists plan designs, deductibles, copays, or benefits (e.g., "$1,000 Deductible", "$20 Specialist Visit Copay"), it is NOT a rate page. For such pages, return an empty array `[]`.
- Premium rate tables specifically list monthly premium costs for coverage tiers (Employee Only/EO/EE, Employee + Spouse/ES, Employee + Children/EC, Employee + Family/EF). Look for tier abbreviations (EO, ES, EC, EF) or full tier names, paired with monthly dollar amounts.
- If a plan name appears on a features comparison page but has no actual monthly premium rate table on that page, do not return any rate entries for it (return an empty list `[]` for that page or plan). Do not hallucinate $0.00 or treat deductibles/copays as premiums.

THREE TYPES OF RATE PAGES — handle all three:

TYPE A – Tier-aggregate tables (most common):
  A table listing one row per coverage tier: Employee Only, Employee + Spouse, Employee + Children, Employee + Family.
  Each tier has a single current and renewal amount.
  → Emit one JSON entry per tier row.

TYPE B – Per-subscriber comparison grids (e.g. "Monthly Premium Comparison Details"):
  A grid listing each enrolled subscriber by name (e.g. "1. AGUILAR VLADIMIRLEE A  FAM  2032.85  2451.35").
  Each row shows THAT individual's current premium ("Current Rate") and new/renewal premium ("New Rate").
  These individual amounts often DIFFER within the same tier because family size varies.
  → CRITICAL: emit one JSON entry per subscriber row using their exact Current Rate and New Rate.
    Do NOT collapse multiple subscribers into one entry per tier — every row must be its own JSON object.
    Use the coverage type column (EMP/ESP/ECH/FAM or EE/ES/EC/FAM) to populate the "tier" field.

TYPE C – Current vs Alternate/Renewal comparison tables (e.g. ADP "Monthly Plan/Contribution Report"):
  These tables have THREE columns side by side: "Current" (left), "Renew My Current Plans" (middle), "Alternate Plans" (right).
  CRITICAL DISTINCTION — the two right columns mean very different things:

  MIDDLE column "Renew My Current Plans":
    → This IS the true renewal of the employee's existing plan. Extract it as:
         plan_name       = the CURRENT plan name (left column header)
         current_monthly = premium from the "Current" column
         renewal_monthly = premium from the "Renew My Current Plans" (middle) column  ← CORRECT RENEWAL

  RIGHT column "Alternate Plans":
    → This is a completely DIFFERENT plan the employee could optionally switch to.
       It has NO "renewal" in the traditional sense — its rate IS its own standalone rate.
    → Extract it as a SEPARATE entry:
         plan_name       = the ALTERNATE plan name (right column header)
         current_monthly = premium from the "Alternate Plans" column
         renewal_monthly = premium from the "Alternate Plans" column  (same value — it is its own rate)
    → DO NOT use the alternate plan's premium as the renewal_monthly for the current plan.

  If there are multiple alternate plans shown on the right, output a separate JSON entry for each.
  If a page only has two columns (Current + Renew), treat the right column as the renewal (no alternate).

TYPE D – Multi-plan side-by-side pages with cross-page splits (CRITICAL — read carefully):

  These PDFs use a 3-column-per-page layout where each page shows three plan columns.
  Each plan column is labelled either "Current Plan" or "Renewal Plan" at the top.

  YEAR-SUFFIX NAMING CONVENTION:
  → Plan names ending in "25 CNT" or "2025" = the CURRENT year plan (effective 2025)
  → Plan names ending in "26 CNT" or "2026" = the RENEWAL year plan (effective 2026)
  → A "25 CNT" plan and its matching "26 CNT" plan ARE THE SAME PLAN — just different contract years.
  → The "Current Plan" column with the "25 CNT" name holds current_monthly amounts.
  → The "Renewal Plan" column with the "26 CNT" name holds renewal_monthly amounts.

  COLUMN PAIRING RULE — always pair columns in LEFT-TO-RIGHT order on each page:
    Column 1 and Column 2 on a page = one (current, renewal) pair for one plan.
    Column 3 on that page = the CURRENT half of the NEXT plan (its renewal is on the NEXT page, column 1).

  WORKED EXAMPLE (exactly matching what you will see):
    Page N — three columns:
      Col 1: "Current Plan"  NY S LBTY NG 30/75/4000/50 EPO 25 CNT → rates: EE=$1,042.14, ES=$2,084.28, EC=$1,771.63, FAM=$2,970.09
      Col 2: "Renewal Plan"  NY S LBTY NG 30/75/4000/50 EPO 26 CNT → rates: EE=$1,184.43, ES=$2,368.87, EC=$2,013.53, FAM=$3,375.64
      Col 3: "Current Plan"  NY S LBTY NG 40/80/3250/60 EPO 25 CNT → rates: EE=$1,059.30, ES=$2,118.60, EC=$1,800.81, FAM=$3,019.00

    Page N+1 — three columns:
      Col 1: "Renewal Plan"  NY S LBTY NG 40/80/3250/60 EPO 26 CNT → rates: EE=$1,204.01, ES=$2,408.02, EC=$2,046.82, FAM=$3,431.43
      Col 2: "Current Plan"  NY S MTRO NG 30/80/3750/60 EPO ME 25 CNT → rates: EE=$970.21, ...
      Col 3: "Renewal Plan"  NY S MTRO NG 30/80/3750/60 EPO ME 26 CNT → rates: EE=$1,120.35, ...

  CORRECT OUTPUT for the above:
    Plan "NY S LBTY NG 30/75/4000/50 EPO 25 CNT":
      → current_monthly=1042.14,  renewal_monthly=1184.43   ← Col 1 current paired with Col 2 renewal (same page)
    Plan "NY S LBTY NG 40/80/3250/60 EPO 25 CNT":
      → current_monthly=1059.30,  renewal_monthly=1204.01   ← Col 3 current (page N) paired with Col 1 renewal (page N+1)
    Plan "NY S MTRO NG 30/80/3750/60 EPO ME 25 CNT":
      → current_monthly=970.21,   renewal_monthly=1120.35   ← Col 2 current paired with Col 3 renewal (same page)

  CRITICAL MISTAKES TO AVOID:
  ✗ WRONG: Pairing NY S LBTY NG 40/80/... 25 CNT (current=1059.30) with the renewal of the PREVIOUS plan (1184.43).
  ✗ WRONG: Treating NY S LBTY NG 40/80/... 26 CNT as a standalone plan with current=1204.01 and renewal=1204.01.
  ✓ CORRECT: Always pair each "25 CNT" current column with its matching "26 CNT" renewal column — whether they are on the same page or on consecutive pages.

  USE THE CURRENT-YEAR PLAN NAME (ending in "25 CNT" or "2025") as the plan_name in the JSON output.
  Do NOT output a separate entry for the renewal-year plan name (e.g., "26 CNT") — it is not a different plan.

COVERAGE TYPE → TIER mapping (for all formats):
  EMP or EE  → "Employee"
  ESP or ES  → "Employee + Spouse"
  ECH or EC  → "Employee + Children"
  FAM or EF  → "Employee + Family"

For each entry, output these fields:
- plan_name: The full plan name shown above or beside the table/grid (e.g., "Anthem Gold PPO NO DED/MOOP 7500")
- tier: One of "Employee", "Employee + Spouse", "Employee + Children", "Employee + Family"
- current_monthly: Current monthly premium amount (number, no $ sign)
- renewal_monthly: New/renewal monthly premium amount (number, no $ sign)
- page: Page number of the document

Return a JSON array of objects with these fields. Return ONLY valid JSON, no markdown formatting.

Here is the layout-preserved extracted text for the pages:
"""
        for page_num, page_text in batch:
            prompt_text += f"\n--- PAGE {page_num} ---\n{page_text}\n"

        try:
            log.info("  Sending pages %d-%d to LLM...", batch[0][0], batch[-1][0])

            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "user", "content": prompt_text}
                ],
                max_tokens=4096,
                temperature=0,
            )
            # Universal Token Monitor
            try:
                import sys as _sys, os as _os
                _cp = _os.path.join(_os.path.dirname(_os.path.abspath(__file__)), '..')
                if _cp not in _sys.path: _sys.path.insert(0, _cp)
                from core.universal_token_monitor import track_usage as _tm
                _page_label = f"pages_{batch[0][0]}_to_{batch[-1][0]}"
                _tm(response.usage, model="gpt-4o", poc_name="renewal-process",
                    file_name="invoice_census", step_name=f"renewal_extraction_{_page_label}")
            except Exception: pass

            result_text = response.choices[0].message.content.strip()
            # Clean up markdown code fences if present
            if result_text.startswith("```"):
                result_text = re.sub(r"^```(?:json)?\s*", "", result_text)
                result_text = re.sub(r"\s*```$", "", result_text)

            try:
                rates_data = json.loads(result_text)
                if isinstance(rates_data, list):
                    for item in rates_data:
                        if all(k in item for k in ["plan_name", "tier", "current_monthly", "renewal_monthly"]):
                            rate = RenewalRate(
                                plan_name=item["plan_name"],
                                tier=item["tier"],
                                current_monthly=float(item["current_monthly"]),
                                renewal_monthly=float(item["renewal_monthly"]),
                                source_page=item.get("page", batch[0][0]),
                            )
                            all_rates.append(rate)
            except json.JSONDecodeError as e:
                log.warning("  Failed to parse LLM response as JSON: %s", e)
                log.warning("  Response: %s", result_text[:500])

        except Exception as e:
            log.error("  LLM API call failed for pages %d-%d: %s",
                      batch[0][0], batch[-1][0], e)

    log.info("  Extracted %d renewal rates from invoice", len(all_rates))
    for rate in all_rates:
        log.info("    %s | %s | Current: $%.2f | Renewal: $%.2f",
                 rate.plan_name, rate.tier, rate.current_monthly, rate.renewal_monthly)

    return all_rates, full_text


# ──────────────────────────────────────────────
# Phase 3 – Match census rows to renewal rates
# ──────────────────────────────────────────────
def get_numbers_without_year(plan_name: str) -> List[int]:
    """Normalize plan name and extract all numbers, excluding year suffixes."""
    p = plan_name.upper().strip()
    p = p.replace(",", "")
    tokens = re.split(r'[^A-Z0-9]', p)
    clean_nums = []
    for token in tokens:
        if not token:
            continue
        digits = re.findall(r'\d+', token)
        for d in digits:
            val = int(d)
            if val in (25, 26, 2025, 2026) and token.endswith(d):
                continue
            clean_nums.append(val)
    return clean_nums


def get_deductible(nums: List[int]) -> int:
    """Get the deductible (largest number >= 500) from a list of plan numbers."""
    large_nums = [n for n in nums if n >= 500]
    if large_nums:
        return max(large_nums)
    return 0


def plans_match_semantically(census_plan: str, rate_plan: str) -> bool:
    """
    Check if a census plan matches a rate plan semantically based on deductible and type keywords.
    """
    c_norm = census_plan.upper()
    r_norm = rate_plan.upper()
    
    # Extract clean numbers (excluding year suffixes)
    c_nums = get_numbers_without_year(c_norm)
    r_nums = get_numbers_without_year(r_norm)
    
    # Check deductible match
    c_ded = get_deductible(c_nums)
    r_ded = get_deductible(r_nums)
    
    if c_ded != r_ded:
        return False
        
    # Check if they share at least one other key number (like copays) or if they share the main name tokens
    common_nums = set(c_nums) & set(r_nums)
    if not common_nums:
        return False
        
    # Check network/type compatibility
    c_types = set()
    if "HDHP" in c_norm: c_types.add("HDHP")
    if "EPO" in c_norm: c_types.add("EPO")
    if "PPO" in c_norm: c_types.add("PPO")
    if "OAPIN" in c_norm: c_types.add("EPO")
    if "OAP" in c_norm and "OAPIN" not in c_norm and "HDHP" not in c_norm: c_types.add("PPO")
    
    r_types = set()
    if "HDHP" in r_norm: r_types.add("HDHP")
    if "EPO" in r_norm: r_types.add("EPO")
    if "PPO" in r_norm: r_types.add("PPO")
    
    if c_types and r_types and not (c_types & r_types):
        return False
        
    # Check carrier/main brand compatibility (e.g. Liberty vs Metro)
    if "LBTY" in c_norm or "LIBERTY" in c_norm:
        if not ("LBTY" in r_norm or "LIBERTY" in r_norm):
            return False
    if "MTRO" in c_norm or "METRO" in c_norm:
        if not ("MTRO" in r_norm or "METRO" in r_norm):
            return False
            
    return True


def match_and_reconcile(
    census_rows: List[CensusRow],
    rates: List[RenewalRate],
) -> List[AuditRecord]:
    """
    Match each census employee to a renewal rate based on
    (plan_name, coverage_tier).
    """
    log.info("Phase 3 – Matching %d census rows to %d renewal rates", len(census_rows), len(rates))

    # Build a lookup: (normalized_plan, tier) -> RenewalRate
    rate_lookup: Dict[Tuple[str, str], RenewalRate] = {}
    for rate in rates:
        # Ignore rates that are zero
        if rate.current_monthly == 0 and rate.renewal_monthly == 0:
            continue
        key = (normalize_plan_name(rate.plan_name), rate.tier)
        if key in rate_lookup:
            existing = rate_lookup[key]
            # Avoid overwriting a valid rate with a zero rate
            if (existing.current_monthly > 0 or existing.renewal_monthly > 0) and (rate.current_monthly == 0 and rate.renewal_monthly == 0):
                continue
            # Prefer rates with non-zero fractional parts (actual premiums) over whole integers (deductibles)
            existing_has_fractional = not existing.current_monthly.is_integer() or not existing.renewal_monthly.is_integer()
            new_has_fractional = not rate.current_monthly.is_integer() or not rate.renewal_monthly.is_integer()
            if existing_has_fractional and not new_has_fractional:
                continue
        rate_lookup[key] = rate
        log.debug("  Rate key: %s", key)

    # Also build plan-only lookup for fallback (match by current premium)
    premium_lookup: Dict[Tuple[str, float], RenewalRate] = {}
    for rate in rates:
        # Ignore rates that are zero
        if rate.current_monthly == 0 and rate.renewal_monthly == 0:
            continue
        key = (normalize_plan_name(rate.plan_name), rate.current_monthly)
        if key in premium_lookup:
            existing = premium_lookup[key]
            existing_has_fractional = not existing.current_monthly.is_integer() or not existing.renewal_monthly.is_integer()
            new_has_fractional = not rate.current_monthly.is_integer() or not rate.renewal_monthly.is_integer()
            if existing_has_fractional and not new_has_fractional:
                continue
        premium_lookup[key] = rate

    records: List[AuditRecord] = []

    for cr in census_rows:
        name = f"{cr.first_name} {cr.last_name}"
        tier = coverage_to_tier(cr.coverage_type, cr.relationship)

        # Skip dependents and waived
        if tier == "DEPENDENT":
            records.append(AuditRecord(
                excel_row=cr.excel_row, ee_row=cr.ee_row,
                employee_name=name, current_plan=cr.current_plan,
                coverage_type=cr.coverage_type, relationship=cr.relationship,
                current_premium=cr.monthly_premium,
                matched_renewal=None, match_method="DEPENDENT",
                status="DEPENDENT", notes="Dependent – premium on primary row",
            ))
            continue

        if tier == "WAIVED":
            records.append(AuditRecord(
                excel_row=cr.excel_row, ee_row=cr.ee_row,
                employee_name=name, current_plan=cr.current_plan,
                coverage_type=cr.coverage_type, relationship=cr.relationship,
                current_premium=cr.monthly_premium,
                matched_renewal=None, match_method="WAIVED",
                status="WAIVED", notes="Employee waived coverage",
            ))
            continue

        if not cr.current_plan:
            records.append(AuditRecord(
                excel_row=cr.excel_row, ee_row=cr.ee_row,
                employee_name=name, current_plan=None,
                coverage_type=cr.coverage_type, relationship=cr.relationship,
                current_premium=cr.monthly_premium,
                matched_renewal=None, match_method="NO_PLAN",
                status="NO_PLAN", notes="No current plan description in census",
            ))
            continue

        norm_plan = normalize_plan_name(cr.current_plan)

        # ── Step 1: Match by current premium amount (most precise – unique per subscriber).
        #   Handles per-subscriber comparison grids (e.g. Anthem "Monthly Premium Comparison
        #   Details") where multiple employees in the same tier have individual rates.
        if cr.monthly_premium:
            premium_key = (norm_plan, cr.monthly_premium)
            premium_rate = premium_lookup.get(premium_key)
            if premium_rate is None:
                # Tolerance scan for floating-point imprecision.
                # Prefer a rate whose plan name matches the census plan name exactly
                # before falling back to any plan with the same premium amount.
                best_premium_rate = None
                for (rp, ramt), rate in premium_lookup.items():
                    if abs(ramt - cr.monthly_premium) <= 0.05:
                        if rp == norm_plan:
                            # Exact plan-name match wins immediately — stop scanning
                            best_premium_rate = rate
                            break
                        elif best_premium_rate is None:
                            # Keep as fallback only if nothing better found yet
                            best_premium_rate = rate
                premium_rate = best_premium_rate


            if premium_rate:
                cr.renewal_amount = premium_rate.renewal_monthly
                records.append(AuditRecord(
                    excel_row=cr.excel_row, ee_row=cr.ee_row,
                    employee_name=name, current_plan=cr.current_plan,
                    coverage_type=cr.coverage_type, relationship=cr.relationship,
                    current_premium=cr.monthly_premium,
                    matched_renewal=premium_rate.renewal_monthly,
                    match_method="PREMIUM_AMOUNT",
                    status="MATCH",
                    notes=f"Matched by current premium ${cr.monthly_premium} → {premium_rate.plan_name} | {premium_rate.tier}",
                ))
                continue

        # ── Step 2: Exact match by (plan_name, tier)
        key = (norm_plan, tier)
        matched_rate = rate_lookup.get(key)

        if matched_rate:
            cr.renewal_amount = matched_rate.renewal_monthly
            records.append(AuditRecord(
                excel_row=cr.excel_row, ee_row=cr.ee_row,
                employee_name=name, current_plan=cr.current_plan,
                coverage_type=cr.coverage_type, relationship=cr.relationship,
                current_premium=cr.monthly_premium,
                matched_renewal=matched_rate.renewal_monthly,
                match_method="EXACT_PLAN_TIER",
                status="MATCH",
                notes=f"Matched to {matched_rate.plan_name} | {matched_rate.tier}",
            ))
            continue

        # ── Step 3: Fuzzy plan name match or semantic match
        best_match = None
        best_score = 0
        for (rp, rt), rate in rate_lookup.items():
            if rt == tier:
                # If they match semantically, select it immediately!
                if plans_match_semantically(cr.current_plan, rate.plan_name):
                    best_match = rate
                    best_score = 1.0
                    break
                score = _plan_similarity(norm_plan, rp)
                if score > best_score:
                    best_score = score
                    best_match = rate

        if best_match and (best_score >= 0.7 or plans_match_semantically(cr.current_plan, best_match.plan_name)):
            cr.renewal_amount = best_match.renewal_monthly
            records.append(AuditRecord(
                excel_row=cr.excel_row, ee_row=cr.ee_row,
                employee_name=name, current_plan=cr.current_plan,
                coverage_type=cr.coverage_type, relationship=cr.relationship,
                current_premium=cr.monthly_premium,
                matched_renewal=best_match.renewal_monthly,
                match_method="FUZZY_PLAN_TIER",
                status="MATCH",
                notes=f"Fuzzy matched to {best_match.plan_name} | {best_match.tier} (score={best_score:.2f})",
            ))
            continue

        # ── Step 4: No match
        records.append(AuditRecord(
            excel_row=cr.excel_row, ee_row=cr.ee_row,
            employee_name=name, current_plan=cr.current_plan,
            coverage_type=cr.coverage_type, relationship=cr.relationship,
            current_premium=cr.monthly_premium,
            matched_renewal=None, match_method="NO_MATCH",
            status="NOT_FOUND",
            notes=f"No matching renewal rate found for plan '{cr.current_plan}' tier '{tier}'",
        ))


    # Summary
    matches = sum(1 for r in records if r.status == "MATCH")
    dependents = sum(1 for r in records if r.status == "DEPENDENT")
    waived = sum(1 for r in records if r.status == "WAIVED")
    no_plan = sum(1 for r in records if r.status == "NO_PLAN")
    not_found = sum(1 for r in records if r.status == "NOT_FOUND")
    log.info("  Results: %d MATCH | %d DEPENDENT | %d WAIVED | %d NO_PLAN | %d NOT_FOUND",
             matches, dependents, waived, no_plan, not_found)

    return records


def _plan_similarity(plan1: str, plan2: str) -> float:
    """Simple word-overlap similarity for plan names."""
    words1 = set(plan1.split())
    words2 = set(plan2.split())
    if not words1 or not words2:
        return 0.0
    intersection = words1 & words2
    union = words1 | words2
    return len(intersection) / len(union)


# ──────────────────────────────────────────────
# Phase 4 – Write renewal amounts to census
# ──────────────────────────────────────────────
def write_renewals(
    census_path: Path,
    out_path: Path,
    sheet: str,
    header_row: int,
    census_rows: List[CensusRow],
    census_type: int = 1,
) -> None:
    log.info("Phase 4 – Writing renewal amounts to %s (Type %d census)", out_path, census_type)

    wb = load_workbook(str(census_path))
    ws = wb[sheet]

    # Find the Renewal column by scanning the header row
    # Type 2 uses "Renewal" as well, so the same scan works for both types.
    renewal_col = None
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and str(val).strip() == "Renewal":
            renewal_col = col
            break

    if renewal_col is None:
        default_col = 9 if census_type == 2 else 13
        log.warning(
            "  'Renewal' column not found at header row %d, defaulting to column %d",
            header_row, default_col,
        )
        renewal_col = default_col

    log.info("  Writing to column %d (Renewal)", renewal_col)

    for cr in census_rows:
        if cr.renewal_amount is not None:
            ws.cell(row=cr.excel_row, column=renewal_col, value=cr.renewal_amount)

    wb.save(str(out_path))
    wb.close()
    log.info("  Saved updated census to %s", out_path)


# ──────────────────────────────────────────────
# Phase 5 – Generate audit log
# ──────────────────────────────────────────────
def write_audit_log(records: List[AuditRecord], rates: List[RenewalRate], out_path: Path) -> None:
    log.info("Phase 5 – Writing audit log to %s", out_path)

    # Audit records
    audit_rows = []
    for r in records:
        audit_rows.append({
            "excel_row": r.excel_row,
            "ee_row": r.ee_row,
            "employee_name": r.employee_name,
            "current_plan": r.current_plan,
            "coverage_type": r.coverage_type,
            "relationship": r.relationship,
            "current_premium": r.current_premium,
            "matched_renewal": r.matched_renewal,
            "match_method": r.match_method,
            "status": r.status,
            "notes": r.notes,
        })

    df_audit = pd.DataFrame(audit_rows)

    # Rate table extracted
    rate_rows = []
    for rate in rates:
        rate_rows.append({
            "plan_name": rate.plan_name,
            "tier": rate.tier,
            "current_monthly": rate.current_monthly,
            "renewal_monthly": rate.renewal_monthly,
            "source_page": rate.source_page,
        })
    df_rates = pd.DataFrame(rate_rows)

    with pd.ExcelWriter(str(out_path), engine="openpyxl") as writer:
        df_audit.to_excel(writer, sheet_name="Audit_All", index=False)

        exceptions = df_audit[~df_audit["status"].isin(["MATCH", "DEPENDENT", "WAIVED"])]
        exceptions.to_excel(writer, sheet_name="Exceptions", index=False)

        if not df_rates.empty:
            df_rates.to_excel(writer, sheet_name="Extracted_Rates", index=False)

        summary = pd.DataFrame([{
            "total_rows": len(records),
            "MATCH": len(df_audit[df_audit["status"] == "MATCH"]),
            "DEPENDENT": len(df_audit[df_audit["status"] == "DEPENDENT"]),
            "WAIVED": len(df_audit[df_audit["status"] == "WAIVED"]),
            "NO_PLAN": len(df_audit[df_audit["status"] == "NO_PLAN"]),
            "NOT_FOUND": len(df_audit[df_audit["status"] == "NOT_FOUND"]),
        }])
        summary.to_excel(writer, sheet_name="Summary", index=False)

    log.info("  Audit log written (%d total rows, %d exceptions)",
             len(records), len(exceptions))


# ──────────────────────────────────────────────
# Phase 6 – Save extracted text
# ──────────────────────────────────────────────
def save_extracted_text(text: str, pdf_name: str, output_dir: Path) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    txt_name = Path(pdf_name).stem + "_extracted.txt"
    out_path = output_dir / txt_name
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(text)
    log.info("Phase 6 – Extracted text saved to %s", out_path)


# ──────────────────────────────────────────────
# Find poppler path automatically
# ──────────────────────────────────────────────
def find_poppler_path() -> Optional[str]:
    """Search common locations for poppler binaries."""
    env_path = os.getenv("POPPLER_PATH")
    if env_path and os.path.exists(env_path):
        return env_path

    # Search common Windows locations
    import glob
    search_paths = [
        os.path.expanduser("~/Downloads/**/poppler*/Library/bin"),
        os.path.expanduser("~/Downloads/**/poppler*/bin"),
        "C:/Program Files/poppler*/Library/bin",
        "C:/Program Files/poppler*/bin",
    ]

    for pattern in search_paths:
        matches = glob.glob(pattern, recursive=True)
        for match in matches:
            if os.path.exists(os.path.join(match, "pdftoppm.exe")):
                return match

    return None


# ──────────────────────────────────────────────
# Pre-flight checks
# ──────────────────────────────────────────────
def preflight(census_path: Path, invoice_path: Path) -> None:
    errors = []
    if not census_path.exists():
        errors.append(f"Census file not found: {census_path}")
    if not invoice_path.exists():
        errors.append(f"Invoice file/directory not found: {invoice_path}")
    if errors:
        for e in errors:
            log.error(e)
        sys.exit(1)
    log.info("Pre-flight checks passed.")


# ──────────────────────────────────────────────
# CLI entry point
# ──────────────────────────────────────────────
def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Invoice-to-Census Renewal Premium Pipeline")
    p.add_argument("--census", required=False, default=None, help="Path to census .xlsx file")
    p.add_argument("--invoices", required=False, default=None, help="Path to invoice PDF file or directory")
    p.add_argument("--out-census", required=False, default=None, help="Path for updated census output .xlsx")
    p.add_argument("--census-sheet", default="Census", help="Sheet name in census workbook (default: Census)")
    p.add_argument("--header-row", type=int, default=21,
                   help="1-based row number of the header in census (default: 21)")
    p.add_argument("--use-llm", action="store_true", default=True,
                   help="Use OpenAI LLM for rate extraction (default: True)")
    p.add_argument("--no-llm", action="store_true",
                   help="Disable LLM, use OCR-only extraction")
    return p.parse_args()


def main() -> None:
    args = parse_args()

    invoice_val = args.invoices
    if not invoice_val:
        invoice_val = input("invoice path here: ").strip()
        if (invoice_val.startswith('"') and invoice_val.endswith('"')) or (invoice_val.startswith("'") and invoice_val.endswith("'")):
            invoice_val = invoice_val[1:-1].strip()

    census_val = args.census
    if not census_val:
        census_val = input("census path here: ").strip()
        if (census_val.startswith('"') and census_val.endswith('"')) or (census_val.startswith("'") and census_val.endswith("'")):
            census_val = census_val[1:-1].strip()

    census_path = Path(census_val)
    invoice_path = Path(invoice_val)

    out_census_val = args.out_census or "output/census_updated.xlsx"

    out_census = Path(out_census_val)

    # Ensure output directory exists
    out_census.parent.mkdir(parents=True, exist_ok=True)

    # Project root for extracted_text folder
    project_root = Path(__file__).parent
    extracted_text_dir = project_root / "extracted_text"

    # ── Pre-flight ──────────────────────────────
    preflight(census_path, invoice_path)

    # ── Find poppler ─────────────────────────────
    poppler_path = find_poppler_path()
    if poppler_path:
        log.info("Using Poppler at: %s", poppler_path)
    else:
        log.warning("Poppler not found. PDF-to-image conversion may fail.")

    # ── Phase 1 ──────────────────────────────────
    # Auto-resolve sheet name if requested sheet is not found
    wb_temp = load_workbook(str(census_path), read_only=True)
    sheet_name = args.census_sheet
    if sheet_name not in wb_temp.sheetnames:
        if "Census" in wb_temp.sheetnames:
            sheet_name = "Census"
        elif "Sheet1" in wb_temp.sheetnames:
            sheet_name = "Sheet1"
        else:
            sheet_name = wb_temp.sheetnames[0]
        log.info("Sheet '%s' not found in census file. Automatically using '%s' instead.", args.census_sheet, sheet_name)
    wb_temp.close()

    # Auto-detect census template type (Type 1 or Type 2)
    wb_detect = load_workbook(str(census_path), data_only=True)
    ws_detect  = wb_detect[sheet_name]
    census_type = detect_census_type(ws_detect)
    wb_detect.close()

    if census_type == 2:
        # Type 2: flat layout, headers at row 1
        header_row = 1
        census_rows = ingest_census_type2(
            census_path,
            sheet=sheet_name,
            header_row=header_row,
        )
    else:
        # Type 1: complex layout. Try to auto-detect the real header row by
        # scanning rows 1-40 for a row that contains known column header keywords.
        # Falls back to the CLI arg (default 21) if nothing is found.
        header_row = args.header_row  # start with CLI default
        wb_scan = load_workbook(str(census_path), data_only=True)
        ws_scan = wb_scan[sheet_name]
        _type1_header_keywords = {
            "first name", "last name", "ee row", "coverage type",
            "current plan description", "monthly total premium",
            "relation-ship to employee", "relationship to employee",
            "date of birth", "renewal",
        }
        for _r in range(1, min(41, ws_scan.max_row + 1)):
            _row_vals = {
                str(ws_scan.cell(row=_r, column=_c).value or "").strip().lower()
                for _c in range(1, ws_scan.max_column + 1)
            }
            _matches = _row_vals & _type1_header_keywords
            if len(_matches) >= 3:  # require at least 3 known header keywords
                header_row = _r
                log.info(
                    "  Auto-detected Type 1 header row at row %d (matched: %s)",
                    header_row, ", ".join(sorted(_matches)),
                )
                break
        else:
            log.info(
                "  Header row auto-detection found nothing; using default row %d",
                header_row,
            )
        wb_scan.close()
        census_rows = ingest_census(
            census_path,
            sheet=sheet_name,
            header_row=header_row,
        )

    # ── Phase 2 ──────────────────────────────────
    # Collect all PDF files
    if invoice_path.is_dir():
        pdf_files = sorted(invoice_path.glob("**/*.pdf"))
    else:
        pdf_files = [invoice_path]

    all_rates: List[RenewalRate] = []
    all_text = ""

    for pdf in pdf_files:
        log.info("Processing invoice: %s", pdf.name)
        use_llm = args.use_llm and not args.no_llm

        if use_llm:
            rates, text = extract_rates_with_llm(pdf, poppler_path=poppler_path)
        else:
            rates, text = extract_rates_with_ocr(pdf, poppler_path=poppler_path)

        # Strip any leading UUID job-id prefix (e.g. "<uuid>_original.pdf" → "original.pdf")
        # so the saved file is named after the original input file, not the server-side temp name.
        _uuid_prefix_re = re.compile(
            r'^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}_',
            re.IGNORECASE,
        )
        display_pdf_name = _uuid_prefix_re.sub('', pdf.name)

        all_rates.extend(rates)
        all_text += f"\n{'='*60}\nFile: {display_pdf_name}\n{'='*60}\n{text}\n"

        # Save extracted text per PDF
        save_extracted_text(text, display_pdf_name, extracted_text_dir)

    if not all_rates:
        log.error("No renewal rates extracted from invoices. Check the PDF files.")
        sys.exit(1)

    # ── Phase 3 ──────────────────────────────────
    records = match_and_reconcile(census_rows, all_rates)

    # ── Phase 4 ──────────────────────────────────
    write_renewals(
        census_path, out_census,
        sheet=sheet_name,
        header_row=header_row,
        census_rows=census_rows,
        census_type=census_type,
    )

    # ── Phase 5 ──────────────────────────────────
    # Save the extracted rates to a JSON file as requested by the user
    rates_json = {}
    for rate in all_rates:
        # Ignore rates that are zero
        if rate.current_monthly == 0 and rate.renewal_monthly == 0:
            continue
        if rate.plan_name not in rates_json:
            rates_json[rate.plan_name] = {}
            
        # Avoid overwriting a valid rate with a zero rate or whole integer if a fractional rate exists
        if rate.tier in rates_json[rate.plan_name]:
            existing = rates_json[rate.plan_name][rate.tier]
            if (existing["current"] > 0 or existing["renewal"] > 0) and (rate.current_monthly == 0 and rate.renewal_monthly == 0):
                continue
            existing_has_fractional = not float(existing["current"]).is_integer() or not float(existing["renewal"]).is_integer()
            new_has_fractional = not rate.current_monthly.is_integer() or not rate.renewal_monthly.is_integer()
            if existing_has_fractional and not new_has_fractional:
                continue
                
        rates_json[rate.plan_name][rate.tier] = {
            "current": rate.current_monthly,
            "renewal": rate.renewal_monthly
        }

    # Save the extracted rates to a job-specific JSON file
    job_suffix = ""
    # Extract the job ID (UUIDv4) from the end of the filename if present
    job_suffix = ""
    match = re.search(r"([a-f0-9]{8}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{4}-[a-f0-9]{12})$", out_census.stem, re.IGNORECASE)
    if match:
        job_suffix = "_" + match.group(1)
    out_rates_json_path = out_census.parent / f"extracted_rates{job_suffix}.json"
    with open(out_rates_json_path, "w", encoding="utf-8") as f:
        json.dump(rates_json, f, indent=4)
    log.info("Saved extracted rates JSON to %s", out_rates_json_path)

    log.info("Pipeline complete. Done!")


if __name__ == "__main__":
    main()
